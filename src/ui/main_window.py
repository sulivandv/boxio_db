"""Interface gráfica principal do Boxio.

Este módulo concentra a composição das telas em PySide6 e os componentes
visuais reutilizáveis do sistema. As regras de negócio ficam em
``StockService``; aqui são tratados eventos da interface, renderização de
tabelas, filtros, modais, tooltips e feedback visual ao usuário.

Cuidados de manutenção:
- Evitar editar dados diretamente nas células; sempre abrir modais estruturados.
- Manter os textos lógicos em Qt.UserRole quando uma célula usa widget customizado.
- Centralizar cores/status nos helpers status_color/status_background/status_pill.
"""

from __future__ import annotations

from pathlib import Path

from PySide6.QtCore import Qt, QDate, QObject, QEvent, QPoint, QTimer, QThread, Signal
from PySide6.QtGui import QColor, QPixmap
from PySide6.QtWidgets import (
    QAbstractItemView, QAbstractSpinBox, QCheckBox, QComboBox, QDialog, QDoubleSpinBox, QFrame,
    QDateEdit, QFormLayout, QGridLayout, QHBoxLayout, QHeaderView, QLabel, QLineEdit,
    QApplication, QFileDialog, QMainWindow, QPushButton, QScrollArea, QSpinBox, QStackedWidget,
    QStyledItemDelegate, QStyleOptionViewItem, QTabWidget, QTableWidget, QTableWidgetItem, QTextEdit, QVBoxLayout, QWidget
)

from src.domain.units import is_fractional, physical_total
from src.services.stock_service import StockService, ACTIVE_PURCHASE_STATUSES, FINAL_PURCHASE_STATUSES
from src.updater.update_checker import check_for_update
from src.updater.update_manager import open_manual_download, prepare_update
from src.updater.update_dialog import UpdateAvailableDialog
from src.core.logger import update_logger

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except Exception:  # openpyxl é carregado apenas no momento da exportação
    Workbook = None
    Font = PatternFill = Alignment = Border = Side = get_column_letter = None

PURPLE = "#8A1CF6"
TEXT = "#111827"
MUTED = "#6B7280"
GREEN = "#22C55E"
RED = "#EF4444"
YELLOW = "#EAB308"
BLUE = "#2563EB"
ROOT_DIR = Path(__file__).resolve().parents[2]


class EllipsisDelegate(QStyledItemDelegate):
    """Delegate global para impedir overflow visual em células com textos longos.

    O texto é sempre recortado dentro da área da célula com reticências e o
    conteúdo completo permanece disponível pelo tooltip já aplicado nos itens.
    """
    def paint(self, painter, option, index):
        opt = QStyleOptionViewItem(option)
        self.initStyleOption(opt, index)
        text = opt.text or ""
        available_width = max(12, opt.rect.width() - 12)
        opt.text = opt.fontMetrics.elidedText(text, Qt.ElideRight, available_width)
        super().paint(painter, opt, index)


class InfoDialog(QDialog):
    """Mensagem compacta e sem barra nativa para ajuda, avisos e erros simples."""
    def __init__(self, title: str, text: str, kind: str = "info", parent=None):
        super().__init__(parent)
        self.setWindowFlags(Qt.Dialog | Qt.FramelessWindowHint)
        self.setObjectName("boxioModal")
        self.setModal(True)
        self.setMinimumWidth(420)
        self.setMaximumWidth(640)
        self._drag_pos = None
        layout = QVBoxLayout(self)
        layout.setContentsMargins(18, 16, 18, 16)
        layout.setSpacing(12)
        top = QHBoxLayout()
        icon = QLabel("ℹ️" if kind != "error" else "✕")
        icon.setObjectName("infoDialogIcon" if kind != "error" else "errorDialogIcon")
        icon.setAlignment(Qt.AlignCenter)
        icon.setFixedSize(34, 34)
        title_label = QLabel(title)
        title_label.setObjectName("modalCaption")
        close = QPushButton("✕")
        close.setObjectName("modalCloseButton")
        close.clicked.connect(self.reject)
        top.addWidget(icon)
        top.addWidget(title_label, 1)
        top.addWidget(close)
        body = QLabel(text)
        body.setWordWrap(True)
        body.setTextInteractionFlags(Qt.TextSelectableByMouse)
        ok = QPushButton("OK")
        ok.setObjectName("primaryButton")
        ok.clicked.connect(self.accept)
        btns = QHBoxLayout()
        btns.addStretch()
        btns.addWidget(ok)
        layout.addLayout(top)
        layout.addWidget(body)
        layout.addLayout(btns)

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            self._drag_pos = event.globalPosition().toPoint() - self.frameGeometry().topLeft()
            event.accept()
        super().mousePressEvent(event)

    def mouseMoveEvent(self, event):
        if self._drag_pos is not None and event.buttons() & Qt.LeftButton:
            self.move(event.globalPosition().toPoint() - self._drag_pos)
            event.accept()
        super().mouseMoveEvent(event)

    def mouseReleaseEvent(self, event):
        self._drag_pos = None
        super().mouseReleaseEvent(event)

    def showEvent(self, event):
        super().showEvent(event)
        try:
            apply_boxio_control_symbols(self)
        except Exception:
            pass


def show_info(parent, title: str, text: str):
    InfoDialog(title, text, "info", parent).exec()


def show_error(parent, title: str, text: str):
    InfoDialog(title, friendly_exception_message(text), "error", parent).exec()


def friendly_exception_message(error: Exception | str) -> str:
    """Remove detalhes técnicos de exceções antes de mostrar ao usuário."""
    raw = str(error)
    low = raw.lower()
    if "invalid input syntax for type uuid" in low or "invalid input syntax" in low:
        return "Não foi possível concluir a operação porque existe um vínculo interno inválido ou desatualizado. Atualize a tela e tente novamente."
    if "foreign key" in low or "violates foreign key constraint" in low:
        return "Não foi possível salvar porque um item vinculado não foi encontrado ou está desatualizado."
    if "ambiguousparameter" in low or "could not determine data type" in low:
        return "Não foi possível validar os dados enviados ao banco. Atualize a tela e tente novamente."
    if "duplicate key" in low or "unique constraint" in low:
        return "Já existe um registro com essas informações."
    if "sql:" in low or "background on this error" in low or "psycopg" in low or "sqlalchemy" in low:
        return "Ocorreu uma falha interna ao comunicar com o banco de dados. Atualize a tela e tente novamente. Se persistir, verifique a conexão com o Neon."
    return raw

HELP_CONTENTS = {
    "form_product": {
        "title": "Cadastro de Produto",
        "descricao": "Este formulário é usado para cadastrar ou editar produtos do estoque. As informações aqui definem como o item será identificado, controlado, comprado e exibido nos relatórios.",
        "quando": "Use ao criar um novo item ou ajustar um cadastro já existente.",
        "como": "Preencha os campos principais na ordem: identificação, unidade, estoque, preços e controles especiais.",
        "exemplos": ["Ex.: Caixa com Máscaras Cirúrgicas.", "Ex.: Resina Composta A2 4g.", "Ex.: Álcool 70% 1 L."],
        "impactos": "Um cadastro bem preenchido melhora filtros, compras, alertas de estoque e relatórios."
    },
    "nome_produto": {
        "title": "Nome do produto",
        "descricao": "Nome principal usado para identificar o item no sistema.",
        "quando": "Preencha sempre que criar ou revisar um item.",
        "como": "Use um nome claro, específico e fácil de localizar. O sistema padroniza a capitalização automaticamente.",
        "exemplos": ["Seringa Descartável 5 ml", "Máscara Cirúrgica Tripla", "Resina Composta A2"],
        "impactos": "Esse nome aparece em buscas, relatórios, compras e movimentações."
    },
    "sku": {
        "title": "SKU",
        "descricao": "Código interno de identificação única do produto.",
        "quando": "Use quando quiser manter um padrão interno de codificação. Se deixar em branco, o sistema pode gerar automaticamente.",
        "como": "Prefira códigos curtos, únicos e consistentes. O sistema converte para maiúsculas e normaliza o formato.",
        "exemplos": ["MASK-CX-001", "RESINA-A2-001", "ALCOOL-1L-002"],
        "impactos": "Facilita busca rápida, integração, conferência e evita duplicidade."
    },
    "categoria": {
        "title": "Categoria",
        "descricao": "Grupo em que o item será classificado no estoque.",
        "quando": "Use para organizar o inventário e separar tipos de materiais.",
        "como": "Selecione a categoria mais adequada. Se ela ainda não existir, crie uma nova de forma padronizada.",
        "exemplos": ["Descartáveis", "Materiais Restauradores", "Medicamentos", "EPIs"],
        "impactos": "Melhora filtros, relatórios e geração de SKU."
    },
    "unidade_medida": {
        "title": "Unidade de medida",
        "descricao": "Define como o item será contado no estoque.",
        "quando": "Sempre que cadastrar um produto novo.",
        "como": "Escolha a unidade que melhor representa o item: un, cx, pc, ml, l, g, kg, cm etc.",
        "exemplos": ["un - Unidade para itens individuais", "cx - Caixa para embalagens fechadas", "ml - Mililitros para líquidos"],
        "impactos": "Controla precisão de quantidade, fracionamento, movimentação e cálculo do físico."
    },
    "tipo_material": {
        "title": "Tipo de material",
        "descricao": "Classificação complementar do item para facilitar padronização e relatórios.",
        "quando": "Use quando desejar identificar a natureza do produto além da categoria.",
        "como": "Informe um tipo simples e consistente, como descartável, líquido, instrumental ou restaurador.",
        "exemplos": ["Descartável", "Líquido", "Instrumental", "Material Restaurador"],
        "impactos": "Ajuda em análises e pode apoiar regras futuras do sistema."
    },
    "quantidade_base": {
        "title": "Quantidade base",
        "descricao": "Indica a quantidade física contida em cada unidade cadastrada do item.",
        "quando": "É importante para itens fracionáveis. Para unidades singulares, normalmente fica fixo em 1.",
        "como": "Se o item for fracionável, informe a medida por registro. Se for unitário, o sistema bloqueia em 1.",
        "exemplos": ["1 caixa = 1 cx", "1 frasco = 500 ml", "1 tubo = 4 g"],
        "impactos": "Afeta o cálculo do físico, a exibição nas tabelas e a consistência das movimentações."
    },
    "estoque_atual": {
        "title": "Estoque atual",
        "descricao": "Quantidade disponível no momento do cadastro ou edição.",
        "quando": "Use para carga inicial ou correção controlada do saldo.",
        "como": "Informe quantas unidades, caixas ou registros físicos existem atualmente.",
        "exemplos": ["12 caixas", "8 frascos", "25 unidades"],
        "impactos": "Define o saldo inicial e influencia alertas, compras e dashboard."
    },
    "estoque_minimo": {
        "title": "Estoque mínimo",
        "descricao": "Quantidade limite usada para alertar necessidade de reposição.",
        "quando": "Use para definir o ponto em que o item precisa de atenção.",
        "como": "Informe o menor saldo seguro para operação.",
        "exemplos": ["2 caixas", "5 unidades", "1 frasco"],
        "impactos": "Controla alertas de estoque baixo e priorização de compras."
    },
    "preco_custo": {
        "title": "Preço de custo",
        "descricao": "Valor pago pela clínica na aquisição do item.",
        "quando": "Use quando quiser controlar custo médio, compras e comparações com fornecedores.",
        "como": "Informe o valor em reais referente à aquisição da unidade cadastrada.",
        "exemplos": ["R$ 10,50 por unidade", "R$ 85,00 por caixa"],
        "impactos": "Ajuda em relatórios de custos e histórico de compras."
    },
    "preco_venda": {
        "title": "Preço de venda",
        "descricao": "Valor de referência para venda, repasse ou análise financeira, quando aplicável.",
        "quando": "Use quando o item também fizer parte de um fluxo comercial.",
        "como": "Informe o valor estimado de venda da unidade cadastrada.",
        "impactos": "Pode ser usado em faturamento, orçamentos e margens futuras."
    },
    "marca": {
        "title": "Marca",
        "descricao": "Fabricante ou marca comercial do produto.",
        "quando": "Use quando for importante diferenciar marcas similares.",
        "como": "Selecione uma marca existente ou cadastre uma nova padronizada.",
        "exemplos": ["3M", "FGM", "Maquira", "Dentsply"],
        "impactos": "Ajuda na busca, comparação comercial e controle de qualidade."
    },
    "fornecedor": {
        "title": "Fornecedor",
        "descricao": "Empresa responsável pelo fornecimento do item.",
        "quando": "Use para associar o produto ao parceiro comercial mais frequente ou principal.",
        "como": "Selecione um fornecedor cadastrado ou crie um novo registro padronizado.",
        "impactos": "Facilita compras, histórico, comparação e devoluções."
    },
    "estoque_local": {
        "title": "Estoque / local",
        "descricao": "Local físico onde o item ficará armazenado.",
        "quando": "Use quando a clínica quiser separar estoque principal, salas, armários ou filiais.",
        "como": "Selecione o local correto para o armazenamento do item.",
        "impactos": "Prepara o sistema para multiestoque e melhora rastreabilidade."
    },
    "descricao": {
        "title": "Descrição",
        "descricao": "Campo livre para registrar observações relevantes sobre o item.",
        "quando": "Use quando precisar complementar o cadastro com detalhes importantes.",
        "como": "Escreva informações objetivas, como composição, uso, apresentação ou cuidados.",
        "impactos": "Ajuda a equipe a entender melhor o item e reduz dúvidas operacionais."
    },
    "controla_lote": {
        "title": "Controla lote",
        "descricao": "Ativa o controle de lote para o produto, permitindo rastrear grupos de fabricação ou fornecimento.",
        "quando": "Use quando o item precisar de rastreabilidade por lote, como alimentos, medicamentos, produtos químicos, materiais hospitalares ou itens com exigência operacional de origem.",
        "como": "Ative esta opção quando cada entrada do produto precisar guardar ou consultar o lote de fabricação. Em cenários futuros, o sistema poderá exigir lote nas movimentações e recebimentos desse item.",
        "exemplos": ["Medicamentos com lote do fabricante.", "Produtos químicos com lote de produção.", "Materiais que precisam de rastreabilidade em caso de problema de qualidade."],
        "impactos": "Melhora rastreabilidade de entrada e saída, ajuda em auditorias, facilita conferência de origem e amplia o controle logístico do estoque.",
        "observacoes": "Recomendado quando a clínica precisa saber de qual lote veio cada produto ou quando existe risco operacional em misturar origens diferentes."
    },
    "controla_validade": {
        "title": "Controla validade",
        "descricao": "Ativa o controle de data de validade do item dentro do estoque.",
        "quando": "Use para produtos perecíveis ou com vencimento, como alimentos, medicamentos, cosméticos, materiais estéreis, resinas e outros itens com prazo limitado de uso.",
        "como": "Ative a opção e selecione a data de validade no calendário. O sistema passa a exigir essa informação no cadastro e poderá utilizá-la em alertas e relatórios.",
        "exemplos": ["Anestésicos com vencimento.", "Materiais estéreis com prazo de uso.", "Resinas e produtos químicos com data de expiração."],
        "impactos": "Permite controle de vencimento, ajuda a aplicar FIFO/FEFO, reduz perdas por expiração e melhora a segurança operacional.",
        "observacoes": "Se esta opção estiver marcada, o campo de validade se torna obrigatório para manter o cadastro consistente."
    },
    "controla_serial": {
        "title": "Controla serial",
        "descricao": "Ativa o rastreamento individual por número de série para cada unidade do item.",
        "quando": "Use para eletrônicos, equipamentos, máquinas, ativos patrimoniais ou qualquer produto em que cada unidade tenha identificação única.",
        "como": "Marque esta opção quando o item precisar ser acompanhado individualmente. Em evoluções futuras, o sistema poderá exigir serial por unidade em entradas, saídas e movimentações específicas.",
        "exemplos": ["Autoclaves", "Equipamentos odontológicos", "Eletrônicos com serial de fábrica", "Ativos rastreáveis da clínica"],
        "impactos": "Permite histórico individual por unidade, ajuda em garantia, manutenção, suporte técnico e controle patrimonial.",
        "observacoes": "Ideal para itens de maior valor ou cuja rastreabilidade individual seja importante para gestão e auditoria."
    },
    "data_validade": {
        "title": "Data de validade",
        "descricao": "Data de vencimento do item.",
        "quando": "É obrigatória quando o controle de validade estiver ativo.",
        "como": "Selecione a data pelo calendário. Evite digitação manual.",
        "impactos": "Usada em alertas, relatórios e prioridade de consumo."
    },
    "movement_form": {
        "title": "Movimentação de estoque",
        "descricao": "Registra entradas, saídas ou ajustes de saldo de um item.",
        "quando": "Use sempre que houver consumo, recebimento, acerto de inventário ou reposição.",
        "impactos": "Toda movimentação alimenta o histórico, a auditoria e o saldo do estoque."
    },
    "mov_tipo": {"title": "Tipo", "descricao": "Define o efeito da movimentação no estoque.", "quando": "Escolha antes de informar a quantidade.", "como": "Entrada soma ao saldo, saída reduz o saldo e ajuste redefine/corrige o estoque.", "exemplos": ["Entrada: recebimento de compra", "Saída: uso em procedimento", "Ajuste: correção após contagem"]},
    "mov_quantidade": {"title": "Quantidade", "descricao": "Quantidade movimentada na operação.", "quando": "Sempre que registrar entrada, saída ou ajuste.", "como": "Informe a quantidade usando a unidade escolhida. Itens fracionáveis aceitam casas decimais.", "impactos": "Altera o saldo do item imediatamente."},
    "mov_unidade": {"title": "Unidade utilizada", "descricao": "Unidade usada nesta movimentação.", "quando": "Use quando o item puder ser movimentado em unidades compatíveis.", "como": "Selecione uma unidade da mesma dimensão do item, como ml, l, g ou kg.", "impactos": "O sistema converte a quantidade quando necessário."},
    "mov_responsavel": {"title": "Responsável", "descricao": "Pessoa responsável pela movimentação.", "quando": "Use para rastrear quem executou a operação.", "impactos": "A informação fica registrada em auditoria e relatórios."},
    "mov_origem": {"title": "Origem / destino", "descricao": "Indica de onde o material veio ou para onde foi destinado.", "exemplos": ["Fornecedor", "Sala 2", "Ajuste de inventário", "Procedimento clínico"]},
    "mov_observacao": {"title": "Observação", "descricao": "Campo livre para explicar a movimentação.", "como": "Registre contexto útil, como motivo do ajuste, número do pedido ou intercorrências."},
    "purchase_request_form": {"title": "Solicitação de compra", "descricao": "Formulário para gerar um pedido interno de reposição ou aquisição.", "quando": "Use quando um item precisar ser comprado ou reposto.", "impactos": "A solicitação passa a compor o fluxo de compras e o acompanhamento operacional."},
    "compra_item": {"title": "Item", "descricao": "Produto que será comprado.", "como": "Selecione o item da lista. O vínculo será usado em compras, recebimentos e atualização do estoque."},
    "compra_quantidade": {"title": "Quantidade solicitada", "descricao": "Quantidade desejada para compra.", "como": "Informe o total necessário respeitando a unidade do item.", "impactos": "Usada para aprovações, pedido e conferência de recebimento."},
    "compra_solicitante": {"title": "Solicitante", "descricao": "Pessoa que solicitou a compra.", "impactos": "Melhora rastreabilidade e comunicação interna."},
    "compra_prioridade": {"title": "Prioridade", "descricao": "Define o grau de urgência da compra.", "como": "Escolha Baixa, Normal, Alta ou Urgente conforme a necessidade operacional.", "impactos": "Ajuda a equipe a decidir a ordem de atendimento."},
    "compra_justificativa": {"title": "Justificativa", "descricao": "Explica por que a compra é necessária.", "como": "Descreva o motivo de forma objetiva, como reposição, aumento de demanda ou item próximo do fim."},
    "compra_observacoes": {"title": "Observações", "descricao": "Campo complementar para informações adicionais da solicitação.", "exemplos": ["Preferência por marca X", "Necessário até sexta-feira", "Conferir lote"]},
    "purchase_manage_form": {"title": "Fluxo de compra", "descricao": "Tela para acompanhar todo o processo da compra, desde a análise até o recebimento ou devolução.", "quando": "Use para atualizar status, fornecedor, pedido, previsão e registrar recebimentos.", "impactos": "Mantém rastreabilidade, histórico e atualização correta do estoque."},
    "compra_status": {"title": "Status", "descricao": "Representa a etapa atual da compra.", "quando": "Atualize sempre que o processo avançar ou sofrer alteração.", "como": "Selecione a etapa que melhor representa a situação atual do pedido.", "exemplos": ["Em Análise: pedido em avaliação.", "Aguardando Aprovação: esperando autorização.", "Aguardando Pedido: aprovado, mas ainda não enviado ao fornecedor.", "Pedido Realizado: compra já emitida ao fornecedor.", "Compra Parcial Recebida: parte do material já chegou.", "Compra Recebida Integralmente: tudo foi entregue.", "Pendente de devolução: houve excedente, avaria ou necessidade de retorno.", "Finalizado: processo encerrado.", "Cancelado/Rejeitado: compra interrompida."], "impactos": "O status orienta a equipe e influencia relatórios operacionais."},
    "compra_responsavel": {"title": "Responsável pela compra", "descricao": "Pessoa que está acompanhando ou executando a compra.", "impactos": "Facilita controle, cobrança e rastreabilidade."},
    "compra_fornecedor": {"title": "Fornecedor", "descricao": "Fornecedor envolvido na compra.", "como": "Selecione a empresa responsável pelo atendimento do pedido.", "impactos": "Usado em histórico comercial, recebimento e devoluções."},
    "compra_numero_pedido": {"title": "Número do pedido", "descricao": "Identificador do pedido emitido ao fornecedor.", "quando": "Use quando a compra já tiver número ou protocolo.", "impactos": "Facilita conferência, busca e auditoria."},
    "compra_valor": {"title": "Valor", "descricao": "Valor total da compra ou do pedido.", "como": "Informe o custo previsto ou confirmado em reais.", "impactos": "Apoia controle financeiro e comparação com fornecedores."},
    "compra_previsao": {"title": "Previsão de entrega", "descricao": "Data estimada para chegada do material.", "quando": "Use após confirmação do pedido pelo fornecedor.", "impactos": "Ajuda no planejamento do estoque e acompanhamento da compra."},
    "compra_obs": {"title": "Observação da alteração", "descricao": "Campo para registrar contexto da última alteração feita no pedido.", "exemplos": ["Fornecedor confirmou envio", "Pedido aguardando aprovação financeira", "Entrega reagendada"]},
    "compra_recebimento": {"title": "Quantidade recebida", "descricao": "Quantidade efetivamente entregue pelo fornecedor nesta confirmação.", "quando": "Use ao receber total ou parcialmente o material.", "como": "Informe somente o que foi conferido fisicamente.", "impactos": "Atualiza o estoque e o andamento do pedido."},
    "compra_avarias": {"title": "Avarias / divergências", "descricao": "Campo para registrar problemas encontrados no recebimento.", "quando": "Use quando houver danos, falta, sobra, troca indevida ou divergência do pedido.", "impactos": "Pode gerar devolução, troca e registro no histórico."},
    "compra_timeline": {"title": "Histórico da compra", "descricao": "Mostra a linha do tempo de alterações e ações do pedido.", "quando": "Consulte para entender todo o histórico da solicitação.", "impactos": "Ajuda em auditoria e rastreabilidade completa."},
    "recent_activity": {"title": "Atividade recente", "descricao": "Exibe um resumo das últimas operações registradas no sistema.", "quando": "Use para acompanhar rapidamente o que mudou no estoque e nas compras.", "impactos": "Facilita monitoramento operacional diário."},
    "low_stock": {"title": "Alerta de estoque baixo", "descricao": "Lista itens com estoque igual ou abaixo do mínimo definido.", "quando": "Use para priorizar reposição e evitar falta de materiais.", "impactos": "Ajuda a equipe a decidir compras com maior urgência."},
    "supplier_comparison_form": {"title": "Comparar fornecedores", "descricao": "Tela para registrar e comparar preços, prazos e histórico de compra entre fornecedores de um mesmo produto.", "quando": "Use ao cotar, comprar ou revisar fornecedores.", "impactos": "Ajuda a identificar melhor preço, melhor prazo e fornecedor mais utilizado."},
    "sup_fornecedor": {"title": "Fornecedor", "descricao": "Fornecedor avaliado ou utilizado nesta cotação/compra.", "como": "Selecione um fornecedor cadastrado ou crie um novo.", "impactos": "Alimenta histórico comercial e comparações futuras."},
    "sup_preco_cotado": {"title": "Preço cotado", "descricao": "Valor informado pelo fornecedor durante a cotação.", "quando": "Use antes da compra ser fechada.", "impactos": "Permite comparar propostas mesmo que a compra não tenha sido concluída."},
    "sup_preco_pago": {"title": "Preço pago", "descricao": "Valor efetivamente pago após a compra.", "quando": "Use quando a compra foi realizada.", "impactos": "Ajuda no custo histórico do produto e negociação futura."},
    "sup_data_cotacao": {"title": "Data da cotação", "descricao": "Data em que o preço foi consultado junto ao fornecedor.", "impactos": "Ajuda a saber se a informação comercial ainda está atualizada."},
    "sup_data_compra": {"title": "Data da compra", "descricao": "Data em que a compra foi realizada.", "impactos": "Ajuda a rastrear compras e histórico financeiro."},
    "sup_prazo": {"title": "Prazo de entrega", "descricao": "Quantidade de dias prevista ou realizada para entrega.", "impactos": "Ajuda a escolher fornecedor por velocidade de atendimento."},
    "sup_status": {"title": "Status da negociação", "descricao": "Etapa comercial do contato com o fornecedor.", "exemplos": ["Cotado", "Em negociação", "Comprado", "Recebido", "Cancelado"]},
    "sup_responsavel": {"title": "Responsável", "descricao": "Pessoa que realizou ou acompanhou a cotação/compra.", "impactos": "Melhora rastreabilidade."},
    "sup_observacao": {"title": "Observação", "descricao": "Anotações úteis sobre preço, condição, prazo ou atendimento."},
    "category_detail": {"title": "Categoria e itens vinculados", "descricao": "Tela para editar a categoria e visualizar produtos associados.", "quando": "Use para revisar organização do inventário.", "impactos": "Alterações na categoria afetam filtros, relatórios e listagem dos produtos."},
    "cat_nome": {"title": "Nome da categoria", "descricao": "Nome usado para agrupar produtos semelhantes.", "exemplos": ["Descartáveis", "Medicamentos", "Materiais Restauradores"]},
    "cat_descricao": {"title": "Descrição da categoria", "descricao": "Explicação opcional sobre o objetivo ou conteúdo da categoria."},
    "registry_page": {"title": "Cadastros e estrutura", "descricao": "Área administrativa para gerenciar categorias, marcas, fornecedores, locais de estoque e responsáveis.", "quando": "Use para manter listas padronizadas e evitar digitação livre.", "impactos": "Esses cadastros alimentam formulários, filtros, relatórios e regras operacionais."},
    "registry_categories": {"title": "Categorias", "descricao": "Classificação dos produtos. Use para organizar o inventário e melhorar relatórios."},
    "registry_brands": {"title": "Marcas", "descricao": "Lista padronizada de marcas/fabricantes para evitar duplicidade de nomes."},
    "registry_suppliers": {"title": "Fornecedores", "descricao": "Base comercial utilizada em compras, cotações, histórico e devoluções."},
    "registry_warehouses": {"title": "Estoques/Locais", "descricao": "Locais físicos de armazenamento, como estoque principal, salas, armários ou unidades."},
    "registry_responsibles": {"title": "Responsáveis", "descricao": "Pessoas vinculadas a movimentações, compras, recebimentos e auditoria."},
    "detail_dialog": {"title": "Detalhamento", "descricao": "Lista filtrada de informações do dashboard ou de uma visão específica.", "quando": "Use para abrir ações diretamente a partir de indicadores, alertas ou relatórios resumidos."},
    "product_actions": {"title": "Ações do item", "descricao": "Menu rápido de operações disponíveis para o produto selecionado.", "quando": "Use para editar, movimentar, comprar, comparar fornecedores ou consultar histórico."}
}


def _help_to_html_list(value):
    if not value:
        return ""
    items = value if isinstance(value, (list, tuple)) else [value]
    lis = "".join(f"<li>{str(item)}</li>" for item in items if str(item).strip())
    return f"<ul style='margin-top:4px; margin-bottom:10px'>{lis}</ul>" if lis else ""


def resolve_help_content(spec):
    if isinstance(spec, dict):
        return spec
    key = str(spec).strip()
    return HELP_CONTENTS.get(key, {"title": "Ajuda", "descricao": key})


def build_help_html(spec) -> tuple[str, str]:
    payload = resolve_help_content(spec)
    title = payload.get("title", "Ajuda")
    parts = []
    for label, key in [
        ("Descrição resumida", "descricao"),
        ("Quando utilizar", "quando"),
        ("Como preencher", "como"),
        ("Exemplos práticos", "exemplos"),
        ("Impacto no sistema", "impactos"),
        ("Observações importantes", "observacoes"),
    ]:
        value = payload.get(key)
        if not value:
            continue
        if isinstance(value, (list, tuple)):
            content = _help_to_html_list(value)
        else:
            content = f"<p style='margin-top:4px; margin-bottom:10px'>{value}</p>"
        parts.append(f"<h3 style='margin:10px 0 2px 0; color:#111827; font-size:11pt'>{label}</h3>{content}")
    if not parts:
        parts.append("<p>Sem conteúdo adicional.</p>")
    html = f"<div style='font-family:Segoe UI; font-size:10pt; color:#111827'><h2 style='margin-top:0; color:#111827'>{title}</h2>{''.join(parts)}</div>"
    return title, html


class HelpContentDialog(QDialog):
    """Janela explicativa mais robusta para as ajudas contextuais do sistema."""
    def __init__(self, spec, parent=None):
        super().__init__(parent)
        self.setWindowFlags(Qt.Dialog | Qt.FramelessWindowHint)
        self.setObjectName("boxioModal")
        self.setModal(True)
        self.resize(620, 520)
        self._drag_pos = None
        title, html = build_help_html(spec)
        layout = QVBoxLayout(self)
        layout.setContentsMargins(18, 16, 18, 16)
        layout.setSpacing(12)
        top = QHBoxLayout()
        icon = QLabel("i")
        icon.setObjectName("infoDialogIcon")
        icon.setAlignment(Qt.AlignCenter)
        icon.setFixedSize(34, 34)
        title_label = QLabel(title)
        title_label.setObjectName("modalCaption")
        close = QPushButton("✕")
        close.setObjectName("modalCloseButton")
        close.clicked.connect(self.reject)
        top.addWidget(icon)
        top.addWidget(title_label, 1)
        top.addWidget(close)
        body = QTextEdit()
        body.setReadOnly(True)
        body.setHtml(html)
        body.setObjectName("helpTextBody")
        body.setMinimumHeight(360)
        ok = QPushButton("Entendi")
        ok.setObjectName("primaryButton")
        ok.clicked.connect(self.accept)
        btns = QHBoxLayout()
        btns.addStretch()
        btns.addWidget(ok)
        layout.addLayout(top)
        layout.addWidget(body, 1)
        layout.addLayout(btns)

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            self._drag_pos = event.globalPosition().toPoint() - self.frameGeometry().topLeft()
            event.accept()
        super().mousePressEvent(event)

    def mouseMoveEvent(self, event):
        if self._drag_pos is not None and event.buttons() & Qt.LeftButton:
            self.move(event.globalPosition().toPoint() - self._drag_pos)
            event.accept()
        super().mouseMoveEvent(event)

    def mouseReleaseEvent(self, event):
        self._drag_pos = None
        super().mouseReleaseEvent(event)


def show_help(parent, spec):
    HelpContentDialog(spec, parent).exec()


def confirm_action(parent, title: str, message: str, details: str = "", confirm_text: str = "Confirmar", cancel_text: str = "Cancelar") -> bool:
    dialog = ActionChoiceDialog(
        title,
        message,
        details,
        [("confirm", confirm_text, "primaryButton"), ("cancel", cancel_text, None)],
        parent,
    )
    return dialog.exec() == QDialog.Accepted and dialog.choice == "confirm"


class ToastNotification(QLabel):
    """Confirmação visual discreta para ações concluídas com sucesso."""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setObjectName("toastSuccess")
        self.setAlignment(Qt.AlignCenter)
        self.setVisible(False)
        self._hide_timer = QTimer(self)
        self._hide_timer.setSingleShot(True)
        self._hide_timer.timeout.connect(self.hide)

    def show_message(self, text: str, duration: int = 2600):
        self.setText(f"✓ {text}")
        self.adjustSize()
        margin = 24
        parent = self.parentWidget()
        if parent:
            self.move(max(margin, parent.width() - self.width() - margin), margin + 44)
        self.raise_()
        self.show()
        self._hide_timer.start(duration)


class HelpIcon(QLabel):
    """Ícone de ajuda contextual padronizado com abertura de ajuda detalhada."""
    def __init__(self, help_spec, parent=None):
        super().__init__("i", parent)
        self.setObjectName("helpIcon")
        self._help_spec = help_spec
        payload = resolve_help_content(help_spec)
        summary = payload.get("descricao") or payload.get("title") or "Ajuda"
        self.setToolTip(summary)
        self.setCursor(Qt.PointingHandCursor)
        self.setAlignment(Qt.AlignCenter)
        self.setFixedSize(14, 14)

    def mousePressEvent(self, event):
        show_help(self, self._help_spec)
        super().mousePressEvent(event)


HEADER_HELP_TEXTS = {
    "Qtd. Base": "Quantidade física contida em cada registro do item. Para unidade, caixa e pacote, o valor fica 1. Para itens fracionáveis, representa o volume/peso/comprimento por unidade cadastrada.",
    "Físico": "Total físico calculado: estoque atual multiplicado pela quantidade base e unidade de medida.",
    "Status": "Indica a prioridade operacional do item: Sem estoque, Estoque baixo ou Em estoque.",
    "Compra": "Mostra se existe uma solicitação de compra ativa vinculada ao item.",
}

def clean_header_text(text: str) -> str:
    """Remove adornos visuais do cabeçalho, preservando o nome lógico da coluna."""
    for marker in ["  🔍", "🔍", "  ⚲", "⚲", "  ⏷", "⏷", "  🔽", "🔽", "  ▼", "▼", "  ▾", "▾", "  ℹ", "ℹ"]:
        text = text.replace(marker, "")
    return text.strip()


def label_with_help(text: str, help_text: str, object_name: str = "formLabel") -> QWidget:
    """Cria um rótulo com ícone ℹ️ ao lado para campos e seções importantes."""
    box = QWidget()
    box.setObjectName("transparentBox")
    layout = QHBoxLayout(box)
    layout.setContentsMargins(0, 0, 0, 0)
    layout.setSpacing(4)
    lab = QLabel(text)
    lab.setObjectName(object_name)
    layout.addWidget(lab)
    layout.addWidget(HelpIcon(help_text))
    layout.addStretch()
    return box


def checkbox_with_help(checkbox: QCheckBox, help_text: str) -> QWidget:
    """Associa um checkbox a um ícone de ajuda discreto."""
    box = QWidget()
    box.setObjectName("transparentBox")
    layout = QHBoxLayout(box)
    layout.setContentsMargins(0, 0, 0, 0)
    layout.setSpacing(4)
    layout.addWidget(checkbox)
    layout.addWidget(HelpIcon(help_text))
    layout.addStretch()
    return box


def money(v) -> str:
    return f"R$ {float(v or 0):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def fmt_num(v) -> str:
    try:
        if float(v).is_integer():
            return str(int(float(v)))
        return f"{float(v):.2f}".replace(".", ",")
    except Exception:
        return str(v)


def status_color(status: str) -> QColor:
    if status in {"Em estoque", "Finalizado", "Compra Recebida Integralmente", "Entrada"}:
        return QColor(GREEN)
    if status in {"Sem estoque", "Cancelado", "Rejeitado", "Saída"}:
        return QColor(RED)
    if status in {"Estoque baixo", "Solicitação Criada", "Em Análise", "Aguardando Aprovação", "Aguardando Pedido", "Pedido Realizado", "Compra Parcial Recebida"}:
        return QColor(YELLOW)
    return QColor(TEXT)


def status_background(status: str) -> QColor:
    if status in {"Em estoque", "Finalizado", "Compra Recebida Integralmente", "Entrada"}:
        return QColor("#DCFCE7")
    if status in {"Sem estoque", "Cancelado", "Rejeitado", "Saída"}:
        return QColor("#FEE2E2")
    if status in {"Estoque baixo", "Solicitação Criada", "Em Análise", "Aguardando Aprovação", "Aguardando Pedido", "Pedido Realizado", "Compra Parcial Recebida"}:
        return QColor("#FEF3C7")
    return QColor("#FFFFFF")


def style_status_item(item: QTableWidgetItem, status: str):
    item.setForeground(status_color(status))
    item.setToolTip(f"Status: {status}")
    return item


def status_pill(status: str) -> QLabel:
    """Cria o marcador visual de status em formato de cápsula, sem pintar a célula inteira."""
    label = QLabel(status or "-")
    fg = status_color(status).name()
    bg = status_background(status).name()
    label.setAlignment(Qt.AlignCenter)
    label.setMinimumHeight(24)
    label.setMaximumHeight(26)
    label.setMinimumWidth(112)
    label.setSizePolicy(label.sizePolicy().horizontalPolicy(), label.sizePolicy().verticalPolicy())
    label.setStyleSheet(
        f"background:{bg}; color:{fg}; border-radius:12px; "
        "padding:5px 12px; font-size:8pt; font-weight:800;"
    )
    label.setToolTip(f"Status: {status}")
    return label


def set_status_cell(table: QTableWidget, row: int, col: int, status: str):
    # A célula mantém o valor lógico em UserRole para filtros/ordenação,
    # mas o texto visível fica exclusivamente no QLabel customizado.
    # Isso elimina o texto duplicado desenhado por baixo da cápsula/pill.
    item = QTableWidgetItem("")
    item.setData(Qt.UserRole, status or "")
    item.setToolTip(f"Status: {status}")
    item.setForeground(QColor("transparent"))
    table.setItem(row, col, item)
    if status:
        wrapper = QWidget()
        wrapper.setStyleSheet("background: transparent;")
        layout = QHBoxLayout(wrapper)
        layout.setContentsMargins(4, 6, 4, 6)
        layout.setSpacing(0)
        layout.addStretch()
        layout.addWidget(status_pill(status))
        layout.addStretch()
        table.setCellWidget(row, col, wrapper)
        table.setRowHeight(row, max(table.rowHeight(row), 58))


def iso_to_qdate(value: str | None) -> QDate:
    if not value:
        return QDate.currentDate()
    for fmt in ("yyyy-MM-dd", "dd/MM/yyyy"):
        d = QDate.fromString(str(value)[:10], fmt)
        if d.isValid():
            return d
    return QDate.currentDate()


def qdate_to_iso(widget: QDateEdit) -> str:
    return widget.date().toString("yyyy-MM-dd")


def configure_date_picker(widget: QDateEdit, tooltip: str = "Selecione a data pelo calendário."):
    widget.setCalendarPopup(True)
    widget.setDisplayFormat("dd/MM/yyyy")
    widget.setDate(QDate.currentDate())
    widget.setMinimumDate(QDate(2000, 1, 1))
    widget.setToolTip(tooltip)
    try:
        widget.lineEdit().setReadOnly(True)
    except Exception:
        pass




class AppDialog(QDialog):
    """Janela modal sem barra padrão do sistema operacional.

    A classe aplica visual moderno/clean para modais e mantém funcionalidades
    essenciais: botão customizado de fechar, arraste pelo cabeçalho e tecla Esc.
    Os formulários herdados chamam ``add_chrome`` após criar o layout principal.
    """
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.setWindowFlags(Qt.Dialog | Qt.FramelessWindowHint)
        self.setObjectName("boxioModal")
        self.setAttribute(Qt.WA_TranslucentBackground, False)
        self._drag_pos = None

    def add_chrome(self, layout: QVBoxLayout, title: str = "") -> None:
        # Não desenha mais uma barra de título visual completa. Mantém apenas
        # o botão de fechar integrado ao conteúdo para preservar o visual clean.
        bar = QFrame(); bar.setObjectName("modalTitleBar")
        h = QHBoxLayout(bar); h.setContentsMargins(0, 0, 0, 0); h.setSpacing(0)
        close = QPushButton("✕"); close.setObjectName("modalCloseButton"); close.setFixedSize(28, 28)
        close.setToolTip("Fechar janela")
        close.clicked.connect(self.reject)
        h.addStretch(1); h.addWidget(close, 0)
        layout.addWidget(bar)
        self._drag_handle = bar

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            self._drag_pos = event.globalPosition().toPoint() - self.frameGeometry().topLeft()
            event.accept()
        super().mousePressEvent(event)

    def mouseMoveEvent(self, event):
        if self._drag_pos is not None and event.buttons() & Qt.LeftButton:
            self.move(event.globalPosition().toPoint() - self._drag_pos)
            event.accept()
        super().mouseMoveEvent(event)

    def mouseReleaseEvent(self, event):
        self._drag_pos = None
        super().mouseReleaseEvent(event)

def table_cell_text(table: QTableWidget, row: int, column: int) -> str:
    """Texto lógico da célula usado para filtros, ordenação auxiliar e tooltips.

    Algumas células, como Status, usam QWidget/QLabel customizado para evitar
    renderização duplicada. Nesses casos o texto visual fica no widget, e o
    valor pesquisável fica em Qt.UserRole.
    """
    item = table.item(row, column)
    if item is not None:
        logical = item.data(Qt.UserRole)
        if logical not in (None, ""):
            return str(logical)
        return item.text() or ""
    widget = table.cellWidget(row, column)
    if widget is not None:
        label = widget.findChild(QLabel)
        if label is not None:
            return label.text()
    return ""




def visible_or_selected_rows(table: QTableWidget) -> list[int]:
    """Retorna as linhas que serão exportadas para Excel.

    Regra de UX: se houver linhas selecionadas, exporta somente a seleção;
    caso contrário, exporta todas as linhas visíveis após filtros rápidos ou
    filtros avançados. Linhas ocultas por filtros não entram no arquivo.
    """
    selected = sorted({idx.row() for idx in table.selectionModel().selectedRows()}) if table.selectionModel() else []
    if selected:
        return [r for r in selected if not table.isRowHidden(r)]
    return [r for r in range(table.rowCount()) if not table.isRowHidden(r)]


def export_table_to_excel(parent: QWidget, table: QTableWidget, default_name: str = "exportacao_tabela.xlsx") -> None:
    """Exporta uma QTableWidget para .xlsx com cabeçalhos e formatação básica.

    A função é genérica e reutilizável por qualquer tela do projeto. Ela coleta
    valores lógicos de células comuns e de células com widgets customizados
    (como as cápsulas de status), respeita filtros/seleção e deixa o usuário
    escolher o local de salvamento.
    """
    if Workbook is None:
        show_error(parent, "Exportação indisponível", "A biblioteca openpyxl não está instalada. Execute: pip install openpyxl")
        return
    rows = visible_or_selected_rows(table)
    visible_cols = [c for c in range(table.columnCount()) if not table.isColumnHidden(c)]
    if not rows or not visible_cols:
        show_info(parent, "Exportar Excel", "Não há dados visíveis ou selecionados para exportar.")
        return
    path, _ = QFileDialog.getSaveFileName(parent, "Salvar arquivo Excel", default_name, "Planilha Excel (*.xlsx)")
    if not path:
        return
    if not path.lower().endswith(".xlsx"):
        path += ".xlsx"
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Exportação"
        headers = []
        for c in visible_cols:
            header = table.horizontalHeaderItem(c)
            text = header.text() if header else f"Coluna {c+1}"
            headers.append(clean_header_text(text))
        ws.append(headers)
        for r in rows:
            ws.append([table_cell_text(table, r, c) for c in visible_cols])
        header_fill = PatternFill("solid", fgColor="F3F4F6")
        header_font = Font(bold=True, color="111827")
        thin = Side(style="thin", color="E5E7EB")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(bottom=thin)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="center")
        for idx, col in enumerate(ws.columns, start=1):
            max_len = 10
            for cell in col:
                max_len = max(max_len, min(len(str(cell.value or "")), 60))
            ws.column_dimensions[get_column_letter(idx)].width = max_len + 2
        ws.freeze_panes = "A2"
        wb.save(path)
        show_info(parent, "Exportação concluída", f"Arquivo Excel salvo com sucesso:\n{path}")
    except Exception as exc:
        show_error(parent, "Erro ao exportar", f"Não foi possível gerar o Excel.\n\nDetalhes: {exc}")


def apply_table_filters(table: QTableWidget):
    """Compatibilidade: os filtros inteligentes por coluna foram removidos.

    A busca/filtro principal das páginas permanece ativa, mas a janela
    "" e seu evento de cabeçalho foram desativados.
    Mantemos esta função como no-op para preservar chamadas existentes sem
    regressão.
    """
    for r in range(table.rowCount()):
        table.setRowHidden(r, False)


class ActionCard(QFrame):
    def __init__(self, title: str, value: str | int, subtitle: str, button_text: str, callback=None, color=PURPLE):
        super().__init__()
        self.setObjectName("card")
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 16, 20, 16)
        layout.setSpacing(7)
        t = QLabel(title); t.setObjectName("cardTitle")
        v = QLabel(str(value)); v.setObjectName("cardValue")
        s = QLabel(subtitle); s.setObjectName("cardSubtitle"); s.setStyleSheet(f"color:{color};")
        b = QPushButton(button_text)
        b.setToolTip(f"Abrir detalhamento operacional de: {title}")
        if callback:
            b.clicked.connect(callback)
        layout.addWidget(t); layout.addWidget(v); layout.addWidget(s); layout.addWidget(b)



# Configuração padrão das tabelas: aplica seleção por linha, cabeçalho interativo,
# altura segura, tooltips e persistência de largura de colunas.
def configure_table(table: QTableWidget, service: StockService | None = None, name: str = ""):
    table.setEditTriggers(QAbstractItemView.NoEditTriggers)
    table.setItemDelegate(EllipsisDelegate(table))
    table.setSelectionBehavior(QAbstractItemView.SelectRows)
    table.setSelectionMode(QAbstractItemView.SingleSelection)
    table.setSortingEnabled(True)
    header = table.horizontalHeader()
    header.setSectionResizeMode(QHeaderView.Interactive)
    header.setStretchLastSection(False)
    header.setMinimumSectionSize(72)
    header.setDefaultAlignment(Qt.AlignLeft | Qt.AlignVCenter)
    table.verticalHeader().setVisible(False)
    table.verticalHeader().setDefaultSectionSize(58)
    table.verticalHeader().setMinimumSectionSize(54)
    table.setWordWrap(False)
    table.setAlternatingRowColors(False)
    table.setTextElideMode(Qt.ElideRight)
    # Cabeçalho limpo: mantém apenas ordenação padrão e ícone ℹ em colunas técnicas.
    for i in range(table.columnCount()):
        item = table.horizontalHeaderItem(i)
        if item:
            raw = clean_header_text(item.text())
            help_text = HEADER_HELP_TEXTS.get(raw)
            item.setText(f"{raw}{'  ℹ' if help_text else ''}")
            item.setToolTip((help_text + "\n\n" if help_text else "") + "Clique no cabeçalho para ordenar. Arraste a borda das colunas para redimensionar.")
    table.setToolTip("Clique no cabeçalho para ordenar. Arraste a borda das colunas para redimensionar.")
    DEFAULT_WIDTHS = {
        "inventory": {0: 330, 1: 95, 2: 160, 3: 150, 4: 130, 5: 80, 6: 105, 7: 95, 8: 105, 9: 110, 10: 150},
        "low_stock": {0: 240, 1: 105, 2: 90, 3: 90, 4: 150, 5: 150},
        "recent_activities": {0: 130, 1: 230, 2: 80, 3: 145, 4: 130, 5: 150},
        "movements": {0: 260, 1: 115},
        "purchases": {0: 260, 1: 115, 2: 165},
    }
    for idx, width in DEFAULT_WIDTHS.get(name, {}).items():
        if idx < table.columnCount():
            table.setColumnWidth(idx, width)
    if service and name:
        widths = service.table_columns(name)
        for i, w in enumerate(widths[:table.columnCount()]):
            floor = DEFAULT_WIDTHS.get(name, {}).get(i, 72)
            table.setColumnWidth(i, max(floor, int(w)))
        table._pending_column_save = None
        def persist_widths(*_):
            # Não grava a cada pixel do arraste. Agenda a gravação para reduzir
            # chamadas remotas ao Neon e deixar o redimensionamento fluido.
            if getattr(table, "_pending_column_save", None):
                table._pending_column_save.stop()
            timer = QTimer(table)
            timer.setSingleShot(True)
            def save():
                try:
                    service.save_table_columns(name, [table.columnWidth(i) for i in range(table.columnCount())])
                except Exception:
                    pass
            timer.timeout.connect(save)
            table._pending_column_save = timer
            timer.start(800)
        header.sectionResized.connect(persist_widths)


def fill_reference_combo(combo: QComboBox, service: StockService, table: str, placeholder: str = "Registro", allow_empty: bool = False):
    combo.blockSignals(True)
    combo.clear()
    combo.addItem(f"+ Criar novo(a) {placeholder.lower()}", "__new__")
    if allow_empty:
        combo.addItem("Nenhum", "")
    for rec in getattr(service, table)():
        combo.addItem(rec.get("nome", ""), rec.get("id"))
    combo.blockSignals(False)


def select_combo_value(combo: QComboBox, value):
    for i in range(combo.count()):
        if combo.itemData(i) == value:
            combo.setCurrentIndex(i)
            return
    if combo.count():
        combo.setCurrentIndex(1 if combo.count() > 1 else 0)


def maybe_create_reference(parent, combo: QComboBox, service: StockService, table: str, label: str, allow_empty: bool = False):
    if combo.currentData() != "__new__":
        return
    dialog = TextInputDialog(f"Novo cadastro: {label}", f"Nome de {label.lower()}:", parent=parent)
    if dialog.exec() == QDialog.Accepted and dialog.input.text().strip():
        try:
            rec = service.create_reference(table, dialog.input.text().strip())
            fill_reference_combo(combo, service, table, label, allow_empty)
            select_combo_value(combo, rec["id"])
        except Exception as e:
            show_error(parent, "Validação", str(e))
            select_combo_value(combo, "" if allow_empty else None)
    else:
        select_combo_value(combo, "" if allow_empty else None)



def apply_boxio_control_symbols(root):
    """Garante símbolos visuais claros nos controles numéricos."""
    for spin in root.findChildren(QSpinBox):
        spin.setButtonSymbols(QAbstractSpinBox.ButtonSymbols.PlusMinus)
    for spin in root.findChildren(QDoubleSpinBox):
        spin.setButtonSymbols(QAbstractSpinBox.ButtonSymbols.PlusMinus)

class TextInputDialog(AppDialog):
    def __init__(self, title, label, value="", description="", parent=None):
        super().__init__(parent)
        self.setWindowTitle(title)
        self.resize(460, 220)
        layout = QVBoxLayout(self)
        self.add_chrome(layout, title)
        layout.addWidget(QLabel(label))
        self.input = QLineEdit(value)
        layout.addWidget(self.input)
        layout.addWidget(QLabel("Descrição / observação"))
        self.description = QTextEdit(description)
        self.description.setFixedHeight(70)
        layout.addWidget(self.description)
        buttons = QHBoxLayout()
        ok = QPushButton("Salvar"); ok.setObjectName("primaryButton")
        cancel = QPushButton("Cancelar")
        ok.clicked.connect(self.accept); cancel.clicked.connect(self.reject)
        buttons.addWidget(ok); buttons.addWidget(cancel); buttons.addStretch()
        layout.addLayout(buttons)



class ActionChoiceDialog(AppDialog):
    """Modal padronizado para escolhas de confirmação e decisão."""
    def __init__(self, title: str, message: str, details: str = "", choices: list[tuple[str, str, str | None]] | None = None, parent=None):
        super().__init__(parent)
        self.choice = None
        self.setWindowTitle(title)
        self.resize(640, 220)
        layout = QVBoxLayout(self)
        self.add_chrome(layout, title)
        title_label = QLabel(title)
        title_label.setObjectName("pageTitle")
        title_label.setWordWrap(True)
        layout.addWidget(title_label)
        body = QLabel(message)
        body.setWordWrap(True)
        body.setTextInteractionFlags(Qt.TextSelectableByMouse)
        layout.addWidget(body)
        if details:
            details_label = QLabel(details)
            details_label.setWordWrap(True)
            details_label.setObjectName("pageSubtitle")
            details_label.setTextInteractionFlags(Qt.TextSelectableByMouse)
            layout.addWidget(details_label)
        buttons = QHBoxLayout()
        for key, label, object_name in (choices or []):
            btn = QPushButton(label)
            if object_name:
                btn.setObjectName(object_name)
            btn.clicked.connect(lambda _=False, value=key: self._set_choice(value))
            buttons.addWidget(btn)
        buttons.addStretch()
        layout.addLayout(buttons)

    def _set_choice(self, value: str):
        self.choice = value
        self.accept()




# Modal de criação/edição de produto. Ele centraliza validações de cadastro,
# comportamento dinâmico por unidade de medida e controle obrigatório de validade.
class ProductDialog(AppDialog):
    def __init__(self, service: StockService, product: dict | None = None, parent=None):
        super().__init__(parent)
        self.service = service
        self.product = product
        self.setWindowTitle("Editar Produto" if product else "Adicionar Produto")
        self.resize(940, 720)
        self.setMinimumSize(760, 540)
        main = QVBoxLayout(self)
        self.add_chrome(main, "Editar Produto" if product else "Adicionar Novo Produto")
        title = QLabel("Editar Produto" if product else "Adicionar Novo Produto"); title.setObjectName("pageTitle")
        subtitle = QLabel("Cadastro estruturado com unidade, categoria, marca, fornecedor e multiestoque padronizados."); subtitle.setObjectName("pageSubtitle")
        title_row=QHBoxLayout(); title_row.addWidget(title); title_row.addWidget(HelpIcon("form_product")); title_row.addStretch(); main.addLayout(title_row); main.addWidget(subtitle)
        form_card = QFrame(); form_card.setObjectName("card")
        form = QGridLayout(form_card); form.setContentsMargins(22,22,22,22); form.setHorizontalSpacing(16); form.setVerticalSpacing(10)
        self.nome = QLineEdit(); self.sku = QLineEdit(); self.sku_preview = QLabel("SKU previsto: --"); self.sku_preview.setObjectName("hint"); self.categoria = QComboBox(); self.unidade = QComboBox(); self.tipo_material = QLineEdit()
        self.quantidade_base = QDoubleSpinBox(); self.quantidade_base.setRange(0, 999999999); self.quantidade_base.setDecimals(3); self.quantidade_base.setSingleStep(0.1)
        self.estoque = QDoubleSpinBox(); self.estoque.setRange(0, 999999999); self.estoque.setDecimals(3)
        self.estoque_minimo = QDoubleSpinBox(); self.estoque_minimo.setRange(0, 999999999); self.estoque_minimo.setDecimals(3)
        self.preco_custo = QDoubleSpinBox(); self.preco_custo.setPrefix("R$ "); self.preco_custo.setRange(0,999999999); self.preco_custo.setDecimals(2)
        self.preco_venda = QDoubleSpinBox(); self.preco_venda.setPrefix("R$ "); self.preco_venda.setRange(0,999999999); self.preco_venda.setDecimals(2)
        self.marca = QComboBox(); self.fornecedor = QComboBox(); self.estoque_local = QComboBox(); self.descricao = QTextEdit(); self.descricao.setFixedHeight(78)
        self.controla_lote = QCheckBox("Controla lote"); self.controla_validade = QCheckBox("Controla validade"); self.controla_serial = QCheckBox("Controla serial")
        self.data_validade = QDateEdit(); configure_date_picker(self.data_validade, "Data de vencimento do item. Obrigatória quando 'Controla validade' estiver ativo.")
        self.data_validade_label = label_with_help("Data de validade *", "data_validade")
        tips = {
            self.nome: "Nome padronizado automaticamente com a primeira letra de cada palavra em maiúscula.",
            self.sku: "Se ficar em branco, o sistema gera um SKU com base na categoria/produto. Quando preenchido, será convertido para maiúsculas.",
            self.categoria: "Selecione uma categoria cadastrada ou crie uma nova opção controlada.",
            self.marca: "Marca padronizada: sem digitação livre para evitar duplicidade.",
            self.fornecedor: "Fornecedor usado em relatórios e solicitações de compra.",
            self.estoque_local: "Local de armazenamento. Base para evolução multiestoque.",
            self.quantidade_base: "Para unidades fracionáveis, informe quantidade física por registro. Para un/cx/pc fica bloqueado em 1.",
            self.estoque: "Quantidade de registros em estoque. Ex.: 8 frascos ou 12 caixas.",
            self.estoque_minimo: "Saldo mínimo que dispara alerta de reposição.",
            self.controla_validade: "Quando ativo, a data de validade passa a ser obrigatória no cadastro.",
            self.data_validade: "Escolha a data de vencimento pelo calendário. Não é necessário digitar.",
        }
        for widget, tip in tips.items(): widget.setToolTip(tip)
        fill_reference_combo(self.categoria, service, "categories", "Categoria")
        fill_reference_combo(self.marca, service, "brands", "Marca", allow_empty=True)
        fill_reference_combo(self.fornecedor, service, "suppliers", "Fornecedor", allow_empty=True)
        fill_reference_combo(self.estoque_local, service, "warehouses", "Estoque/local")
        for u in service.units():
            if u.get("ativo", True): self.unidade.addItem(f"{u['codigo']} - {u['descricao']}", u["codigo"])
        field_help = {
            "Nome do produto *": "nome_produto", "SKU": "sku", "Categoria *": "categoria", "Unidade de medida *": "unidade_medida",
            "Tipo de material": "tipo_material", "Quantidade base": "quantidade_base", "Estoque atual": "estoque_atual", "Estoque mínimo": "estoque_minimo",
            "Preço de custo": "preco_custo", "Preço de venda": "preco_venda", "Marca": "marca", "Fornecedor": "fornecedor", "Estoque/local": "estoque_local"
        }
        fields = [
            ("Nome do produto *", self.nome,0,0),("SKU", self.sku,0,1),("Categoria *", self.categoria,1,0),("Unidade de medida *", self.unidade,1,1),
            ("Tipo de material", self.tipo_material,2,0),("Quantidade base", self.quantidade_base,2,1),("Estoque atual", self.estoque,3,0),("Estoque mínimo", self.estoque_minimo,3,1),
            ("Preço de custo", self.preco_custo,4,0),("Preço de venda", self.preco_venda,4,1),("Marca", self.marca,5,0),("Fornecedor", self.fornecedor,5,1),("Estoque/local", self.estoque_local,6,0)
        ]
        for label, widget, r, c in fields:
            box = QVBoxLayout(); box.addWidget(label_with_help(label, field_help.get(label, label))); box.addWidget(widget); form.addLayout(box,r,c)
        desc_box = QVBoxLayout(); desc_box.addWidget(label_with_help("Descrição", "descricao")); desc_box.addWidget(self.descricao); form.addLayout(desc_box,7,0,1,2)
        checks = QHBoxLayout(); checks.setSpacing(14); checks.addWidget(checkbox_with_help(self.controla_lote, "controla_lote")); checks.addWidget(checkbox_with_help(self.controla_validade, "controla_validade")); checks.addWidget(checkbox_with_help(self.controla_serial, "controla_serial")); checks.addStretch(); form.addLayout(checks,8,0,1,2)
        validade_box = QVBoxLayout(); validade_box.addWidget(self.data_validade_label); validade_box.addWidget(self.data_validade); form.addLayout(validade_box,9,0,1,1)
        self.rule_label = QLabel(""); self.rule_label.setObjectName("hint"); form.addWidget(self.sku_preview,10,0,1,2); form.addWidget(self.rule_label,11,0,1,2)
        buttons = QHBoxLayout(); save = QPushButton("Salvar Produto"); save.setObjectName("primaryButton"); cancel = QPushButton("Cancelar"); save.clicked.connect(self.save); cancel.clicked.connect(self.reject); buttons.addWidget(save); buttons.addWidget(cancel); buttons.addStretch(); form.addLayout(buttons,12,0,1,2)
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        scroll.setWidget(form_card)
        main.addWidget(scroll, 1)
        self.unidade.currentIndexChanged.connect(self.apply_unit_rules)
        self.nome.textChanged.connect(self.update_sku_preview)
        self.sku.textChanged.connect(self.update_sku_preview)
        self.categoria.currentIndexChanged.connect(self.update_sku_preview)
        self.marca.currentIndexChanged.connect(self.update_sku_preview)
        self.categoria.currentIndexChanged.connect(lambda: maybe_create_reference(self,self.categoria,service,"categories","Categoria"))
        self.marca.currentIndexChanged.connect(lambda: maybe_create_reference(self,self.marca,service,"brands","Marca",True))
        self.fornecedor.currentIndexChanged.connect(lambda: maybe_create_reference(self,self.fornecedor,service,"suppliers","Fornecedor",True))
        self.estoque_local.currentIndexChanged.connect(lambda: maybe_create_reference(self,self.estoque_local,service,"warehouses","Estoque/local"))
        self.controla_validade.toggled.connect(self.update_validity_visibility)
        self.load_product(); self.apply_unit_rules(); self.update_validity_visibility(); self.update_sku_preview()

    def set_combo(self, combo, value): select_combo_value(combo, value)

    def load_product(self):
        p = self.product or {}
        self.nome.setText(p.get("nome", "")); self.sku.setText(p.get("sku", "")); self.set_combo(self.categoria,p.get("categoria_id")); self.set_combo(self.unidade,p.get("unidade_medida","un"))
        self.tipo_material.setText(p.get("tipo_material", "")); self.quantidade_base.setValue(float(p.get("quantidade_base", 1))); self.estoque.setValue(float(p.get("estoque_atual", 0))); self.estoque_minimo.setValue(float(p.get("estoque_minimo", 0)))
        self.preco_custo.setValue(float(p.get("preco_custo", 0))); self.preco_venda.setValue(float(p.get("preco_venda", 0))); self.set_combo(self.marca,p.get("marca_id","")); self.set_combo(self.fornecedor,p.get("fornecedor_id","")); self.set_combo(self.estoque_local,p.get("multiestoque_id","estoque-principal"))
        self.descricao.setPlainText(p.get("descricao", "")); self.controla_lote.setChecked(bool(p.get("controla_lote", False))); self.controla_validade.setChecked(bool(p.get("controla_validade", False))); self.controla_serial.setChecked(bool(p.get("controla_serial", False)))
        self.data_validade.setDate(iso_to_qdate(p.get("data_validade") or p.get("validade")))

    def apply_unit_rules(self):
        code = self.unidade.currentData(); unit = self.service.unit(code); frac = is_fractional(code, self.service.units())
        self.tipo_material.setText(unit.get("tipo_material_padrao", ""))
        if frac:
            precision = int(unit.get("precisao_decimal", 2)); self.quantidade_base.setEnabled(True); self.quantidade_base.setDecimals(precision)
            if self.quantidade_base.value() <= 0: self.quantidade_base.setValue(float(unit.get("quantidade_base_padrao", 1.0)))
            self.rule_label.setText(f"Unidade fracionável: informe a quantidade física por registro em {code}. Ex.: 500 ml por frasco."); self.estoque.setDecimals(3); self.estoque_minimo.setDecimals(3)
        else:
            self.quantidade_base.setValue(1); self.quantidade_base.setEnabled(False); self.rule_label.setText("Unidade singular: quantidade base bloqueada em 1. Cada registro representa uma unidade completa."); self.estoque.setDecimals(0); self.estoque_minimo.setDecimals(0)

    def update_validity_visibility(self):
        active = self.controla_validade.isChecked()
        self.data_validade.setVisible(active)
        self.data_validade_label.setVisible(active)

    def update_sku_preview(self):
        """Mostra ao usuário a prévia do SKU antes de salvar o produto."""
        manual = self.sku.text().strip()
        if manual:
            preview = manual.upper().replace(" ", "-")
            self.sku_preview.setText(f"SKU informado será padronizado como: {preview}")
        else:
            preview = self.service.preview_sku(self.categoria.currentData() or "", self.nome.text(), self.marca.currentData() or "", self.tipo_material.text())
            self.sku_preview.setText(f"SKU previsto: {preview}")

    def save(self):
        if not self.nome.text().strip(): show_info(self,"Validação","Informe o nome do produto."); return
        if self.controla_validade.isChecked() and not self.data_validade.date().isValid(): show_info(self,"Validação","Informe a data de validade pelo calendário."); return
        if self.categoria.currentData() in (None,"__new__"): show_info(self,"Validação","Selecione uma categoria válida."); return
        payload = {"nome":self.nome.text(),"sku":self.sku.text(),"categoria_id":self.categoria.currentData(),"unidade_medida":self.unidade.currentData(),"tipo_material":self.tipo_material.text(),"quantidade_base":self.quantidade_base.value(),"estoque_atual":self.estoque.value(),"estoque_minimo":self.estoque_minimo.value(),"preco_custo":self.preco_custo.value(),"preco_venda":self.preco_venda.value(),"marca_id":self.marca.currentData() or "","fornecedor_id":self.fornecedor.currentData() or "","descricao":self.descricao.toPlainText(),"controla_lote":self.controla_lote.isChecked(),"controla_validade":self.controla_validade.isChecked(),"data_validade":qdate_to_iso(self.data_validade) if self.controla_validade.isChecked() else "","controla_serial":self.controla_serial.isChecked(),"multiestoque_id":self.estoque_local.currentData()}
        try:
            if self.product: self.service.update_product(self.product["id"], payload)
            else: self.service.add_product(payload)
            self.accept()
        except Exception as e: show_error(self,"Erro",str(e))


class MovementDialog(AppDialog):
    def __init__(self, service: StockService, product: dict, parent=None):
        super().__init__(parent); self.service=service; self.product=product; self.setWindowTitle("Movimentar Estoque"); self.resize(560,430)
        layout=QVBoxLayout(self); self.add_chrome(layout,"Movimentar Estoque"); title_row=QHBoxLayout(); title=QLabel(f"Movimentação: {product['nome']}"); title.setObjectName("pageTitle"); title_row.addWidget(title); title_row.addWidget(HelpIcon("movement_form")); title_row.addStretch(); layout.addLayout(title_row)
        form=QFormLayout(); self.tipo=QComboBox(); self.tipo.addItems(["entrada","saida","ajuste"]); self.quantidade=QDoubleSpinBox(); self.quantidade.setRange(0.001,9999999); self.quantidade.setDecimals(3)
        self.unidade=QComboBox(); product_unit=product["unidade_medida"]; dim=service.unit(product_unit).get("dimensao")
        for u in service.units():
            if u.get("ativo", True) and u.get("dimensao")==dim: self.unidade.addItem(f"{u['codigo']} - {u['descricao']}", u["codigo"])
        select_combo_value(self.unidade, product_unit)
        self.responsavel=QComboBox(); fill_reference_combo(self.responsavel, service, "responsibles", "Responsável"); select_combo_value(self.responsavel,"resp-padrao"); self.responsavel.currentIndexChanged.connect(lambda: maybe_create_reference(self,self.responsavel,service,"responsibles","Responsável"))
        self.origem=QLineEdit(); self.obs=QTextEdit(); self.obs.setFixedHeight(70)
        self.tipo.setToolTip("Entrada soma ao estoque, saída subtrai e ajuste define o saldo final informado."); self.quantidade.setToolTip("Quantidade movimentada na unidade selecionada. O sistema converte quando compatível."); self.unidade.currentIndexChanged.connect(self.apply_unit_rules); self.apply_unit_rules()
        form.addRow(label_with_help("Tipo", "mov_tipo"),self.tipo); form.addRow(label_with_help("Quantidade", "mov_quantidade"),self.quantidade); form.addRow(label_with_help("Unidade utilizada", "mov_unidade"),self.unidade); form.addRow(label_with_help("Responsável", "mov_responsavel"),self.responsavel); form.addRow(label_with_help("Origem/Destino", "mov_origem"),self.origem); form.addRow(label_with_help("Observação", "mov_observacao"),self.obs); layout.addLayout(form)
        hint=QLabel("A operação registra saldo anterior, saldo restante, responsável, data/hora e auditoria."); hint.setObjectName("hint"); layout.addWidget(hint)
        buttons=QHBoxLayout(); save=QPushButton("Registrar"); save.setObjectName("primaryButton"); cancel=QPushButton("Cancelar"); save.clicked.connect(self.save); cancel.clicked.connect(self.reject); buttons.addWidget(save); buttons.addWidget(cancel); buttons.addStretch(); layout.addLayout(buttons)
    def apply_unit_rules(self):
        code = self.unidade.currentData() or self.product.get("unidade_medida", "un")
        self.quantidade.setDecimals(3 if is_fractional(code, self.service.units()) else 0)

    def save(self):
        try:
            self.service.add_movement(self.product["id"], self.tipo.currentText(), self.quantidade.value(), self.unidade.currentData(), self.responsavel.currentData(), self.origem.text(), self.obs.toPlainText()); self.accept()
        except Exception as e: show_error(self,"Erro",str(e))


class PurchaseRequestDialog(AppDialog):
    def __init__(self, service: StockService, product: dict | None = None, parent=None):
        super().__init__(parent); self.service=service; self.product=product; self.setWindowTitle("Solicitar Compra"); self.resize(620,430)
        layout=QVBoxLayout(self); self.add_chrome(layout,"Solicitação de Compra"); title_row=QHBoxLayout(); title=QLabel("Solicitação de Compra"); title.setObjectName("pageTitle"); title_row.addWidget(title); title_row.addWidget(HelpIcon("purchase_request_form")); title_row.addStretch(); layout.addLayout(title_row)
        form=QFormLayout(); self.produto=QComboBox()
        for p in service.products(): self.produto.addItem(f"{p.get('nome')} | {p.get('sku')} | estoque {fmt_num(p.get('estoque_atual'))}", p["id"])
        if product: select_combo_value(self.produto, product["id"])
        self.quantidade=QDoubleSpinBox(); self.quantidade.setRange(0.001,9999999); self.quantidade.setDecimals(3); self.quantidade.setValue(max(float(product.get("estoque_minimo",1) if product else 1),1))
        self.solicitante=QComboBox(); fill_reference_combo(self.solicitante,service,"responsibles","Responsável"); select_combo_value(self.solicitante,"resp-padrao"); self.solicitante.currentIndexChanged.connect(lambda: maybe_create_reference(self,self.solicitante,service,"responsibles","Responsável"))
        self.prioridade=QComboBox(); self.prioridade.addItems(["Baixa","Normal","Alta","Urgente"])
        self.justificativa=QTextEdit(); self.justificativa.setFixedHeight(65); self.obs=QTextEdit(); self.obs.setFixedHeight(65)
        self.produto.setToolTip("Produto que precisa de reposição. A solicitação ficará vinculada a ele."); self.quantidade.setToolTip("Quantidade solicitada na unidade de estoque do produto."); self.produto.currentIndexChanged.connect(self.apply_product_rules); self.apply_product_rules()
        form.addRow(label_with_help("Item", "compra_item"),self.produto); form.addRow(label_with_help("Quantidade solicitada", "compra_quantidade"),self.quantidade); form.addRow(label_with_help("Solicitante", "compra_solicitante"),self.solicitante); form.addRow(label_with_help("Prioridade", "compra_prioridade"),self.prioridade); form.addRow(label_with_help("Justificativa", "compra_justificativa"),self.justificativa); form.addRow(label_with_help("Observações", "compra_observacoes"),self.obs); layout.addLayout(form)
        buttons=QHBoxLayout(); save=QPushButton("Criar Solicitação"); save.setObjectName("primaryButton"); cancel=QPushButton("Cancelar"); save.clicked.connect(self.save); cancel.clicked.connect(self.reject); buttons.addWidget(save); buttons.addWidget(cancel); buttons.addStretch(); layout.addLayout(buttons)
    def apply_product_rules(self):
        p = self.service.get_product(self.produto.currentData())
        if p:
            self.quantidade.setDecimals(3 if is_fractional(p.get("unidade_medida", "un"), self.service.units()) else 0)

    def save(self):
        try:
            self.service.create_purchase_request(self.produto.currentData(), self.quantidade.value(), self.solicitante.currentData(), self.prioridade.currentText(), self.justificativa.toPlainText(), self.obs.toPlainText()); self.accept()
        except Exception as e: show_error(self,"Erro",str(e))


# Modal do fluxo de compras: gerencia status, fornecedor, pedido e recebimento,
# mantendo timeline/auditoria e atualização automática do estoque.
class PurchaseManageDialog(AppDialog):
    STATUSES = ["Em Análise","Aguardando Aprovação","Aguardando Pedido","Pedido Realizado","Compra Parcial Recebida","Compra Recebida Integralmente","Pendente de devolução","Aguardando coleta do fornecedor","Item devolvido","Troca recebida","Finalizado","Cancelado","Rejeitado"]
    def __init__(self, service: StockService, request: dict, parent=None):
        super().__init__(parent); self.service=service; self.request=request; self.setWindowTitle("Fluxo de Compra"); self.resize(850,650)
        layout=QVBoxLayout(self); self.add_chrome(layout,"Fluxo de Compra"); title_row=QHBoxLayout(); title=QLabel(f"Compra: {request.get('produto_nome','')}"); title.setObjectName("pageTitle"); title_row.addWidget(title); title_row.addWidget(HelpIcon("purchase_manage_form")); title_row.addStretch(); layout.addLayout(title_row)
        self.meta=QLabel(f"Status atual: {request.get('status')} | Solicitado: {fmt_num(request.get('quantidade_solicitada'))} {request.get('unidade_medida')} | Recebido: {fmt_num(request.get('quantidade_recebida'))}"); self.meta.setObjectName("pageSubtitle"); layout.addWidget(self.meta)
        form=QGridLayout(); self.status=QComboBox(); self.status.addItems(self.STATUSES); select_combo_value(self.status, request.get("status"))
        self.responsavel=QComboBox(); fill_reference_combo(self.responsavel,service,"responsibles","Responsável",allow_empty=True); select_combo_value(self.responsavel,request.get("responsavel_compra_id",""))
        self.fornecedor=QComboBox(); fill_reference_combo(self.fornecedor,service,"suppliers","Fornecedor",allow_empty=True); select_combo_value(self.fornecedor,request.get("fornecedor_id",""))
        self.numero=QLineEdit(request.get("numero_pedido","")); self.valor=QDoubleSpinBox(); self.valor.setPrefix("R$ "); self.valor.setRange(0,999999999); self.valor.setDecimals(2); self.valor.setValue(float(request.get("valor",0) or 0)); self.previsao=QDateEdit(); configure_date_picker(self.previsao, "Previsão de entrega selecionada por calendário."); self.previsao.setDate(iso_to_qdate(request.get("previsao_entrega"))); self.obs=QTextEdit(); self.obs.setFixedHeight(60)
        widget_help = {"Status":"compra_status", "Responsável pela compra":"compra_responsavel", "Fornecedor":"compra_fornecedor", "Número do pedido":"compra_numero_pedido", "Valor":"compra_valor", "Previsão de entrega":"compra_previsao"}
        widgets=[("Status",self.status,0,0),("Responsável pela compra",self.responsavel,0,1),("Fornecedor",self.fornecedor,1,0),("Número do pedido",self.numero,1,1),("Valor",self.valor,2,0),("Previsão de entrega",self.previsao,2,1)]
        for label,w,r,c in widgets:
            box=QVBoxLayout(); box.addWidget(label_with_help(label, widget_help.get(label, label))); box.addWidget(w); form.addLayout(box,r,c)
        box=QVBoxLayout(); box.addWidget(label_with_help("Observação da alteração", "compra_obs")); box.addWidget(self.obs); form.addLayout(box,3,0,1,2); layout.addLayout(form)
        receive_card=QFrame(); receive_card.setObjectName("card"); rv=QVBoxLayout(receive_card); receive_header=QHBoxLayout(); receive_header.addWidget(QLabel("Recebimento de material")); receive_header.addWidget(HelpIcon("compra_recebimento")); receive_header.addStretch(); rv.addLayout(receive_header); rf=QHBoxLayout(); self.received=QDoubleSpinBox(); self.received.setRange(0,999999999); self.received.setDecimals(3); self.avarias=QLineEdit(); self.received.setToolTip("Quantidade efetivamente recebida. Ao confirmar, o sistema gera entrada automática no estoque."); rf.addWidget(label_with_help("Quantidade recebida", "compra_recebimento")); rf.addWidget(self.received); rf.addWidget(label_with_help("Avarias/divergências", "compra_avarias")); rf.addWidget(self.avarias); self.receive_btn=QPushButton("Confirmar Recebimento"); self.receive_btn.setObjectName("primaryButton"); self.receive_btn.clicked.connect(self.receive); rf.addWidget(self.receive_btn); rv.addLayout(rf); self.receive_feedback=QLabel(""); self.receive_feedback.setObjectName("hint"); rv.addWidget(self.receive_feedback); layout.addWidget(receive_card); self._receiving=False
        timeline_header=QHBoxLayout(); timeline_header.addWidget(QLabel("Histórico da compra")); timeline_header.addWidget(HelpIcon("compra_timeline")); timeline_header.addStretch(); layout.addLayout(timeline_header)
        self.timeline=QTableWidget(0,5); self.timeline.setHorizontalHeaderLabels(["Data","Status","Ação","Responsável","Observação"]); configure_table(self.timeline,service,"purchase_timeline"); layout.addWidget(self.timeline,1); self.refresh_timeline()
        buttons=QHBoxLayout(); save=QPushButton("Salvar Status"); save.setObjectName("primaryButton"); close=QPushButton("Fechar"); save.clicked.connect(self.save_status); close.clicked.connect(self.accept); buttons.addWidget(save); buttons.addWidget(close); buttons.addStretch(); layout.addLayout(buttons)
    def refresh_timeline(self):
        self.timeline.setRowCount(0)
        for ev in self.request.get("timeline", []):
            r=self.timeline.rowCount(); self.timeline.insertRow(r); vals=[ev.get("criado_em",""),ev.get("status",""),ev.get("acao",""),ev.get("responsavel",""),ev.get("observacao","")]
            for c,v in enumerate(vals): item=QTableWidgetItem(str(v)); item.setForeground(QColor(TEXT)); item.setToolTip(str(v)); self.timeline.setItem(r,c,item)
    def save_status(self):
        try:
            self.request=self.service.update_purchase_status(self.request["id"], self.status.currentText(), self.responsavel.currentData() or "", self.fornecedor.currentData() or "", self.numero.text(), self.valor.value(), qdate_to_iso(self.previsao), self.obs.toPlainText()); self.refresh_timeline(); self.operation_success = "Status da compra atualizado com sucesso."
        except Exception as e: show_error(self,"Erro",str(e))
    def receive(self):
        if getattr(self, "_receiving", False):
            return
        qty = float(self.received.value() or 0)
        if qty <= 0:
            show_error(self, "Validação", "Informe uma quantidade recebida maior que zero.")
            return

        requested = float(self.request.get("quantidade_solicitada", 0) or 0)
        already = float(self.request.get("quantidade_recebida", 0) or 0)
        remaining = max(0.0, requested - already)
        overage_action = "stock"
        return_reason = ""

        if qty > remaining and remaining >= 0:
            choice_dialog = ActionChoiceDialog(
                "Quantidade acima do pedido",
                "Foi identificada uma quantidade recebida acima do previsto no pedido. O que deseja fazer?",
                f"Solicitado: {fmt_num(requested)} | Já recebido: {fmt_num(already)} | Restante previsto: {fmt_num(remaining)} | Informado agora: {fmt_num(qty)}",
                [
                    ("stock", "Adicionar excedente ao estoque", None),
                    ("return", "Solicitar devolução ao fornecedor", None),
                    ("cancel", "Cancelar para revisar", None),
                ],
                self,
            )
            if choice_dialog.exec() != QDialog.Accepted or choice_dialog.choice == "cancel":
                return
            if choice_dialog.choice == "return":
                overage_action = "return"
                reason_dialog = TextInputDialog(
                    "Motivo da devolução/troca",
                    "Informe o motivo principal da devolução ou troca com o fornecedor:",
                    value=self.avarias.text(),
                    description="Descreva a divergência encontrada, como excesso recebido, avaria, produto incorreto ou troca necessária.",
                    parent=self,
                )
                if reason_dialog.exec() != QDialog.Accepted:
                    return
                return_reason = (reason_dialog.input.text().strip() or reason_dialog.description.toPlainText().strip())
                if not return_reason:
                    show_error(self, "Validação", "Informe o motivo da devolução ou troca para continuar.")
                    return
            else:
                overage_action = "stock"

        self._receiving = True
        self.receive_btn.setEnabled(False)
        old_text = self.receive_btn.text()
        self.receive_btn.setText("Processando...")
        self.receive_feedback.setText("Confirmando recebimento e atualizando estoque...")
        QApplication.processEvents()
        try:
            self.request = self.service.receive_purchase(self.request["id"], qty, self.responsavel.currentData() or "", self.obs.toPlainText(), self.avarias.text(), overage_action=overage_action, return_reason=return_reason)
            self.refresh_timeline()
            if hasattr(self, "meta"):
                self.meta.setText(f"Status atual: {self.request.get('status')} | Solicitado: {fmt_num(self.request.get('quantidade_solicitada'))} {self.request.get('unidade_medida')} | Recebido: {fmt_num(self.request.get('quantidade_recebida'))}")
            self.received.setValue(0)
            self.receive_feedback.setText("Recebimento confirmado com sucesso.")
            if overage_action == "return" and qty > remaining:
                self.operation_success = "Recebimento confirmado e devolução/troca registrada para o excedente."
            else:
                self.operation_success = "Recebimento confirmado e estoque atualizado com sucesso."
        except Exception as e:
            self.receive_feedback.setText("Não foi possível confirmar o recebimento.")
            show_error(self,"Erro",friendly_exception_message(e))
        finally:
            self._receiving = False
            self.receive_btn.setEnabled(True)
            self.receive_btn.setText(old_text)


class DetailDialog(AppDialog):
    def __init__(self, title: str, service: StockService, rows: list[dict], kind: str, parent=None):
        super().__init__(parent); self.service=service; self.rows=rows; self.kind=kind; self.setWindowTitle(title); self.resize(850,540)
        layout=QVBoxLayout(self); self.add_chrome(layout,title); hrow=QHBoxLayout(); h=QLabel(title); h.setObjectName("pageTitle"); hrow.addWidget(h); hrow.addWidget(HelpIcon("detail_dialog")); hrow.addStretch(); layout.addLayout(hrow)
        if kind == "movements": headers=["Produto","SKU","Tipo","Qtd.","Saldo","Responsável","Data","ID"]
        elif kind == "purchases": headers=["Item","SKU","Status","Qtd. Solicitada","Recebido","Prioridade","Atualizado em","ID"]
        else: headers=["Produto","SKU","Categoria","Estoque","Mínimo","Compra","ID"]
        self.table=QTableWidget(0,len(headers)); self.table.setHorizontalHeaderLabels(headers); self.table.setColumnHidden(len(headers)-1,True); configure_table(self.table,service,f"detail_{kind}"); self.table.itemDoubleClicked.connect(lambda *_: self.open_actions()); layout.addWidget(self.table,1)
        btns=QHBoxLayout(); open_btn=QPushButton("Abrir ações"); close=QPushButton("Fechar"); open_btn.setObjectName("primaryButton"); open_btn.clicked.connect(self.open_actions); close.clicked.connect(self.accept); btns.addWidget(open_btn); btns.addWidget(close); btns.addStretch(); layout.addLayout(btns); self.refresh()
    def refresh(self):
        self.table.setSortingEnabled(False); self.table.setRowCount(0)
        for row in self.rows:
            r=self.table.rowCount(); self.table.insertRow(r)
            if self.kind=="movements": vals=[row.get("produto_nome",""),row.get("sku",""),row.get("tipo",""),fmt_num(row.get("quantidade",0)),fmt_num(row.get("saldo_restante",0)),row.get("responsavel",""),row.get("criado_em",""),row.get("produto_id","")]
            elif self.kind=="purchases": vals=[row.get("produto_nome",""),row.get("sku",""),row.get("status",""),fmt_num(row.get("quantidade_solicitada",0)),fmt_num(row.get("quantidade_recebida",0)),row.get("prioridade",""),row.get("atualizado_em",""),row.get("id","")]
            else:
                vals=[row.get("nome",""),row.get("sku",""),self.service.category_name(row.get("categoria_id","")),fmt_num(row.get("estoque_atual",0)),fmt_num(row.get("estoque_minimo",0)),self.service.purchase_status_for_product(row.get("id")),row.get("id","")]
            for c,v in enumerate(vals):
                if c in {2,5} and str(v): set_status_cell(self.table,r,c,str(v)); continue
                item=QTableWidgetItem(str(v)); item.setForeground(QColor(TEXT)); item.setToolTip(str(v)); self.table.setItem(r,c,item)
        self.table.setSortingEnabled(True)
    def open_actions(self):
        rows=self.table.selectionModel().selectedRows();
        if not rows: return
        hidden=self.table.columnCount()-1; ident=self.table.item(rows[0].row(),hidden).text()
        if self.kind=="purchases":
            req=next((r for r in self.service.purchase_requests() if r["id"]==ident),None)
            if req: PurchaseManageDialog(self.service,req,self).exec(); self.accept()
        else:
            p=self.service.get_product(ident)
            if p: ProductActionDialog(self.service,p,self).exec(); self.accept()


class ProductActionDialog(AppDialog):
    def __init__(self, service: StockService, product: dict, parent=None):
        super().__init__(parent); self.service=service; self.product=product; self.setWindowTitle("Ações do item"); self.resize(420,280)
        layout=QVBoxLayout(self); self.add_chrome(layout,"Ações do Item"); title_row=QHBoxLayout(); title=QLabel(product.get("nome","")); title.setObjectName("pageTitle"); title_row.addWidget(title); title_row.addWidget(HelpIcon("product_actions")); title_row.addStretch(); layout.addLayout(title_row); layout.addWidget(QLabel(f"SKU: {product.get('sku','')} | Estoque: {fmt_num(product.get('estoque_atual',0))} {product.get('unidade_medida','')}"))
        for text, cb, tip in [
            ("Editar item", self.edit, "Abre a edição completa do cadastro."), ("Movimentar estoque", self.move, "Registra entrada, saída ou ajuste."), ("Solicitar compra", self.purchase, "Gera solicitação de reposição vinculada ao item."), ("Comparar fornecedores", self.suppliers, "Mostra histórico de preços, cotações e compras por fornecedor."), ("Visualizar histórico", self.history, "Mostra compras e movimentações relacionadas."),
        ]:
            b=QPushButton(text); b.setToolTip(tip); b.clicked.connect(cb); layout.addWidget(b)
    def edit(self):
        if ProductDialog(self.service,self.product,self).exec()==QDialog.Accepted: self.accept()
    def move(self):
        if MovementDialog(self.service,self.product,self).exec()==QDialog.Accepted: self.accept()
    def purchase(self):
        if PurchaseRequestDialog(self.service,self.product,self).exec()==QDialog.Accepted: self.accept()
    def suppliers(self):
        SupplierComparisonDialog(self.service,self.product,self).exec()

    def history(self):
        reqs=self.service.product_purchase_history(self.product["id"]); movs=[m for m in self.service.recent_movements(500) if m.get("produto_id")==self.product["id"]]
        text="COMPRAS:\n"+"\n".join([f"{r.get('criado_em')} - {r.get('status')} - solicitado {fmt_num(r.get('quantidade_solicitada'))}" for r in reqs]) or "Nenhuma"
        text += "\n\nMOVIMENTAÇÕES:\n"+"\n".join([f"{m.get('criado_em')} - {m.get('tipo')} - {fmt_num(m.get('quantidade'))} - saldo {fmt_num(m.get('saldo_restante'))}" for m in movs])
        show_info(self,"Histórico", text if len(text)<3500 else text[:3500]+"...")


class SupplierComparisonDialog(AppDialog):
    """Modal de comparação comercial entre fornecedores de um mesmo item."""
    def __init__(self, service: StockService, product: dict, parent=None):
        super().__init__(parent); self.service=service; self.product=product; self.setWindowTitle("Comparar Fornecedores"); self.resize(980,620)
        layout=QVBoxLayout(self); self.add_chrome(layout,"Comparar Fornecedores")
        title_row=QHBoxLayout(); title=QLabel(product.get("nome", "")); title.setObjectName("pageTitle"); title_row.addWidget(title); title_row.addWidget(HelpIcon("supplier_comparison_form")); title_row.addStretch(); layout.addLayout(title_row)
        self.summary=QLabel(""); self.summary.setObjectName("pageSubtitle"); layout.addWidget(self.summary)
        self.table=QTableWidget(0,10); self.table.setHorizontalHeaderLabels(["Fornecedor","Preço cotado","Preço pago","Data cotação","Data compra","Prazo","Status","Responsável","Observação","ID"]); self.table.setColumnHidden(9,True); configure_table(self.table,service,"supplier_comparison"); layout.addWidget(self.table,1)
        form=QGridLayout(); self.fornecedor=QComboBox(); fill_reference_combo(self.fornecedor,service,"suppliers","Fornecedor",allow_empty=True); self.fornecedor.currentIndexChanged.connect(lambda: maybe_create_reference(self,self.fornecedor,service,"suppliers","Fornecedor",True))
        self.cotado=QDoubleSpinBox(); self.cotado.setPrefix("R$ "); self.cotado.setRange(0,999999999); self.cotado.setDecimals(2)
        self.pago=QDoubleSpinBox(); self.pago.setPrefix("R$ "); self.pago.setRange(0,999999999); self.pago.setDecimals(2)
        self.data_cotacao=QDateEdit(); configure_date_picker(self.data_cotacao,"Data da cotação")
        self.data_compra=QDateEdit(); configure_date_picker(self.data_compra,"Data da compra")
        self.prazo=QSpinBox(); self.prazo.setRange(0,9999); self.prazo.setSuffix(" dias")
        self.status=QComboBox(); self.status.addItems(["Cotado","Em negociação","Comprado","Recebido","Cancelado"])
        self.responsavel=QComboBox(); fill_reference_combo(self.responsavel,service,"responsibles","Responsável",allow_empty=True)
        self.obs=QLineEdit(); self.obs.setPlaceholderText("Observações comerciais")
        sup_help = {"Fornecedor":"sup_fornecedor", "Preço cotado":"sup_preco_cotado", "Preço pago":"sup_preco_pago", "Data cotação":"sup_data_cotacao", "Data compra":"sup_data_compra", "Prazo":"sup_prazo", "Status":"sup_status", "Responsável":"sup_responsavel", "Observação":"sup_observacao"}
        widgets=[("Fornecedor",self.fornecedor,0,0),("Preço cotado",self.cotado,0,1),("Preço pago",self.pago,0,2),("Data cotação",self.data_cotacao,1,0),("Data compra",self.data_compra,1,1),("Prazo",self.prazo,1,2),("Status",self.status,2,0),("Responsável",self.responsavel,2,1),("Observação",self.obs,2,2)]
        for label,w,r,c in widgets:
            box=QVBoxLayout(); box.addWidget(label_with_help(label, sup_help.get(label, label))); box.addWidget(w); form.addLayout(box,r,c)
        layout.addLayout(form)
        buttons=QHBoxLayout(); add=QPushButton("Adicionar histórico"); add.setObjectName("primaryButton"); add.clicked.connect(self.add_history); close=QPushButton("Fechar"); close.clicked.connect(self.accept); buttons.addWidget(add); buttons.addWidget(close); buttons.addStretch(); layout.addLayout(buttons); self.refresh()
    def refresh(self):
        comp=self.service.supplier_comparison_for_product(self.product["id"]); rows=comp["rows"]
        best=comp.get("best_price") or {}; deadline=comp.get("best_deadline") or {}; used=comp.get("most_used",("",0))
        self.summary.setText(f"Menor preço: {best.get('fornecedor','-')} {money(best.get('preco_pago') or best.get('preco_cotado') or 0)} | Melhor prazo: {deadline.get('fornecedor','-')} {deadline.get('prazo_entrega_dias','-')} dias | Mais usado: {used[0] or '-'} ({used[1]} registros)")
        self.table.setSortingEnabled(False); self.table.setRowCount(0)
        for h in rows:
            vals=[h.get("fornecedor",""), money(h.get("preco_cotado",0)), money(h.get("preco_pago",0)), h.get("data_cotacao",""), h.get("data_compra",""), str(h.get("prazo_entrega_dias",0)), h.get("status_negociacao",""), h.get("responsavel",""), h.get("observacao",""), h.get("id","")]
            r=self.table.rowCount(); self.table.insertRow(r)
            for c,v in enumerate(vals):
                if c==6 and v: set_status_cell(self.table,r,c,v); continue
                item=QTableWidgetItem(str(v)); item.setForeground(QColor(TEXT)); item.setToolTip(str(v)); self.table.setItem(r,c,item)
        self.table.setSortingEnabled(True)
    def add_history(self):
        try:
            self.service.add_supplier_history(self.product["id"], self.fornecedor.currentData() or "", self.cotado.value(), self.pago.value(), qdate_to_iso(self.data_cotacao), qdate_to_iso(self.data_compra), self.prazo.value(), self.status.currentText(), self.responsavel.currentData() or "", self.obs.text())
            self.refresh(); show_info(self,"Fornecedor","Histórico comercial registrado.")
        except Exception as e: show_error(self,"Validação",friendly_exception_message(e))


class CategoryDetailDialog(AppDialog):
    def __init__(self, service: StockService, category: dict, parent=None):
        super().__init__(parent); self.service=service; self.category=category; self.setWindowTitle("Detalhes da Categoria"); self.resize(850,560)
        layout=QVBoxLayout(self); title_row=QHBoxLayout(); title=QLabel("Categoria e itens vinculados"); title.setObjectName("pageTitle"); title_row.addWidget(title); title_row.addWidget(HelpIcon("category_detail")); title_row.addStretch(); layout.addLayout(title_row)
        form=QFormLayout(); self.nome=QLineEdit(category.get("nome","")); self.desc=QTextEdit(category.get("descricao","")); self.desc.setFixedHeight(60); form.addRow(label_with_help("Nome", "cat_nome"),self.nome); form.addRow(label_with_help("Descrição", "cat_descricao"),self.desc); layout.addLayout(form)
        self.table=QTableWidget(0,7); self.table.setHorizontalHeaderLabels(["Produto","SKU","Estoque","Mínimo","Unidade","Status compra","ID"]); self.table.setColumnHidden(6,True); configure_table(self.table,service,"category_items"); self.table.itemDoubleClicked.connect(lambda *_: self.open_item()); layout.addWidget(self.table,1)
        buttons=QHBoxLayout(); save=QPushButton("Salvar Categoria"); save.setObjectName("primaryButton"); open_btn=QPushButton("Abrir item"); close=QPushButton("Fechar"); save.clicked.connect(self.save); open_btn.clicked.connect(self.open_item); close.clicked.connect(self.accept); buttons.addWidget(save); buttons.addWidget(open_btn); buttons.addWidget(close); buttons.addStretch(); layout.addLayout(buttons); self.refresh()
    def refresh(self):
        self.table.setRowCount(0)
        for p in self.service.products_by_category(self.category["id"]):
            r=self.table.rowCount(); self.table.insertRow(r); vals=[p.get("nome",""),p.get("sku",""),fmt_num(p.get("estoque_atual",0)),fmt_num(p.get("estoque_minimo",0)),p.get("unidade_medida",""),self.service.purchase_status_for_product(p["id"]),p["id"]]
            for c,v in enumerate(vals):
                if c==5 and str(v): set_status_cell(self.table,r,c,str(v)); continue
                item=QTableWidgetItem(str(v)); item.setForeground(QColor(TEXT)); item.setToolTip(str(v)); self.table.setItem(r,c,item)
    def save(self):
        try: self.category=self.service.update_reference("categories", self.category["id"], self.nome.text(), self.desc.toPlainText()); show_info(self,"Categoria","Categoria atualizada.")
        except Exception as e: show_error(self,"Validação",friendly_exception_message(e))
    def open_item(self):
        rows=self.table.selectionModel().selectedRows();
        if not rows: return
        pid=self.table.item(rows[0].row(),6).text(); p=self.service.get_product(pid)
        if p: ProductActionDialog(self.service,p,self).exec(); self.refresh()


class UpdateCheckThread(QThread):
    """Verifica atualização online sem travar a interface PySide.

    O app desktop continua responsivo enquanto consulta o endpoint remoto.
    Sinais Qt comunicam sucesso, ausência de atualização ou erro para a janela.
    """
    update_available = Signal(object)
    update_error = Signal(str)

    def run(self):
        try:
            manifest = check_for_update()
            if manifest:
                self.update_available.emit(manifest)
        except Exception as exc:
            update_logger().warning("Falha ao verificar atualização: %s", exc)
            self.update_error.emit(str(exc))


class DataRefreshThread(QThread):
    """Sincroniza cache remoto sem bloquear a navegação da interface."""
    finished_ok = Signal()
    refresh_error = Signal(str)

    def __init__(self, service, parent=None):
        super().__init__(parent)
        self.service = service

    def run(self):
        try:
            if hasattr(self.service, "force_refresh"):
                self.service.force_refresh()
            elif hasattr(self.service, "sync_remote_cache"):
                self.service.sync_remote_cache(force=True)
            self.finished_ok.emit()
        except Exception as exc:
            self.refresh_error.emit(str(exc))


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__(); self.service=StockService(); self.setWindowTitle("Boxio"); self.setWindowFlags(Qt.FramelessWindowHint | Qt.Window); self.resize(1280,780); self._main_drag_pos=None; self._update_thread=None; self._refresh_thread=None; self.setup_ui(); self.refresh_current_page(); QTimer.singleShot(250, self.start_background_sync); QTimer.singleShot(2500, self.check_updates_silent)


    def check_updates_silent(self):
        """Consulta atualizações no início sem interromper o usuário em caso de erro."""
        self._start_update_check(show_no_update=False)

    def check_updates_manual(self):
        """Consulta atualizações a partir do botão do menu lateral."""
        self._start_update_check(show_no_update=True)

    def _start_update_check(self, show_no_update: bool = False):
        self._show_no_update = show_no_update
        self._update_thread = UpdateCheckThread(self)
        self._update_thread.update_available.connect(self.show_update_dialog)
        self._update_thread.update_error.connect(lambda msg: show_error(self, "Atualizações", f"Não foi possível verificar atualizações agora.\n{msg}") if show_no_update else None)
        self._update_thread.finished.connect(lambda: show_info(self, "Atualizações", "Seu sistema já está na versão mais recente.") if show_no_update and not getattr(self, "_update_found", False) else None)
        self._update_found = False
        self._update_thread.start()

    def show_update_dialog(self, manifest):
        """Exibe o modal de atualização disponível dentro do aplicativo."""
        self._update_found = True
        dialog = UpdateAvailableDialog(manifest, self)
        dialog.manual_btn.clicked.connect(lambda: (open_manual_download(manifest), dialog.accept()))
        def prepare():
            try:
                package = prepare_update(manifest)
                show_info(self, "Atualização preparada", f"Pacote baixado e validado com sucesso:\n{package}\n\nFeche o sistema e execute o instalador/updater para concluir.")
                dialog.accept()
            except Exception as exc:
                show_error(self, "Erro na atualização", str(exc))
        dialog.auto_btn.clicked.connect(prepare)
        dialog.exec()


    def eventFilter(self, obj, event):
        if obj is getattr(self, "_main_titlebar", None):
            if event.type() == QEvent.MouseButtonDblClick and event.button() == Qt.LeftButton:
                self.toggle_maximize_restore()
                event.accept(); return True
            if event.type() == QEvent.MouseButtonPress and event.button() == Qt.LeftButton:
                self._main_drag_pos = event.globalPosition().toPoint() - self.frameGeometry().topLeft()
                event.accept(); return True
            if event.type() == QEvent.MouseMove and self._main_drag_pos is not None and event.buttons() & Qt.LeftButton:
                # Se a janela estiver maximizada, o arraste não deve tentar mover
                # a geometria, pois isso causa comportamento inconsistente no Windows.
                if not self.isMaximized():
                    self.move(event.globalPosition().toPoint() - self._main_drag_pos)
                event.accept(); return True
            if event.type() == QEvent.MouseButtonRelease:
                self._main_drag_pos = None
        return super().eventFilter(obj, event)

    def toggle_maximize_restore(self):
        """Alterna a janela principal entre maximizada e restaurada.

        Como a barra nativa do sistema operacional foi removida, este método
        substitui o comportamento do botão padrão de maximizar/restaurar e
        também atualiza o ícone textual para comunicar o estado atual.
        """
        if self.isMaximized():
            self.showNormal()
        else:
            self.showMaximized()
        self.update_maximize_button_icon()

    def update_maximize_button_icon(self):
        """Atualiza ícone e tooltip do botão customizado de maximizar."""
        btn = getattr(self, "_maximize_button", None)
        if not btn:
            return
        if self.isMaximized():
            btn.setText("❐")
            btn.setToolTip("Restaurar janela")
        else:
            btn.setText("□")
            btn.setToolTip("Maximizar janela")

    def changeEvent(self, event):
        super().changeEvent(event)
        if event.type() == QEvent.WindowStateChange:
            self.update_maximize_button_icon()

    def setup_ui(self):
        icon_dir = ROOT_DIR / "assets" / "icons"
        combo_arrow_icon = (icon_dir / "chevron-down.svg").as_uri() if (icon_dir / "chevron-down.svg").exists() else ""
        plus_icon = (icon_dir / "plus.svg").as_uri() if (icon_dir / "plus.svg").exists() else ""
        minus_icon = (icon_dir / "minus.svg").as_uri() if (icon_dir / "minus.svg").exists() else ""
        self.setStyleSheet("""
            QWidget { background: #F8FAFC; color: #111827; font-family: 'Segoe UI'; font-size: 10pt; }
            #sidebar { background: #FFFFFF; border-right: 1px solid #E5E7EB; }
            #brand { font-size: 13pt; font-weight: 800; color: #8A1CF6; }
            #logoBox { background: #F3E8FF; color: #8A1CF6; border-radius: 10px; font-size: 14pt; font-weight: 900; min-width: 34px; min-height: 34px; max-width: 34px; max-height: 34px; }
            #brandContainer { background: transparent; }
            #navButton { text-align: left; padding: 11px 14px; border-radius: 10px; border: none; color: #374151; background: transparent; }
            #navButton:hover { background: #F3E8FF; color: #7E22CE; }
            #primaryButton { background: #8A1CF6; color: white; border: none; border-radius: 12px; padding: 10px 16px; font-weight: 700; }
            #primaryButton:hover { background: #7E22CE; }
            #primaryButton:pressed { background: #6B21A8; }
            QPushButton { background: #FFFFFF; border: 1px solid #E5E7EB; border-radius: 12px; padding: 9px 14px; min-height: 30px; }
            QPushButton:hover { border-color: #8A1CF6; background: #F8F5FF; }
            QPushButton:pressed { background: #EDE9FE; }
            QPushButton:disabled { color: #9CA3AF; background: #F3F4F6; border-color: #E5E7EB; }
            #dangerButton { color: #EF4444; }
            #card { background: #FFFFFF; border: 1px solid #E5E7EB; border-radius: 14px; }
            #boxioModal { background: #FFFFFF; border: 1px solid #D1D5DB; border-radius: 16px; }
            #pageTitle { font-size: 18pt; font-weight: 800; color: #111827; }
            #pageSubtitle { color: #6B7280; }
            #cardTitle { color: #6B7280; font-size: 10pt; }
            #cardValue { font-size: 20pt; font-weight: 800; color: #111827; }
            #cardSubtitle { font-weight: 700; }
            #formLabel { font-size: 9pt; font-weight: 700; color: #374151; }
            #hint { color: #6B7280; font-size: 9pt; }
            #helpIcon { background: rgba(59,130,246,0.10); color: rgba(37,99,235,0.60); border-radius: 7px; font-size: 7pt; font-weight: 700; }
            #helpIcon:hover { background: rgba(59,130,246,0.18); color: rgba(29,78,216,0.85); }
            #helpTextBody { background: #FFFFFF; border: 1px solid #E5E7EB; border-radius: 12px; padding: 8px; }
            #transparentBox { background: transparent; border: none; }
            QLineEdit, QTextEdit, QComboBox, QSpinBox, QDoubleSpinBox { background: #FFFFFF; border: 1px solid #E5E7EB; border-radius: 10px; padding: 8px; selection-background-color: #E9D5FF; selection-color: #111827; }
            QComboBox, QSpinBox, QDoubleSpinBox { min-height: 34px; }
            QComboBox::drop-down { subcontrol-origin: padding; subcontrol-position: top right; width: 28px; border-left: 1px solid #E5E7EB; background: #F8FAFC; border-top-right-radius: 10px; border-bottom-right-radius: 10px; margin: 1px; }
            QComboBox::down-arrow { width: 10px; height: 10px; }
            QComboBox::drop-down:hover { background: #F3F4F6; }
            QComboBox::drop-down:pressed { background: #EDE9FE; }
            QSpinBox::up-button, QDoubleSpinBox::up-button { subcontrol-origin: border; subcontrol-position: top right; width: 24px; background: #F8FAFC; border-left: 1px solid #E5E7EB; border-top-right-radius: 10px; margin: 1px 1px 0 0; }
            QSpinBox::down-button, QDoubleSpinBox::down-button { subcontrol-origin: border; subcontrol-position: bottom right; width: 24px; background: #F8FAFC; border-left: 1px solid #E5E7EB; border-top: 1px solid #E5E7EB; border-bottom-right-radius: 10px; margin: 0 1px 1px 0; }
            QSpinBox::up-arrow, QDoubleSpinBox::up-arrow { width: 12px; height: 12px; }
            QSpinBox::down-arrow, QDoubleSpinBox::down-arrow { width: 12px; height: 12px; }
            QSpinBox::up-button:hover, QDoubleSpinBox::up-button:hover, QSpinBox::down-button:hover, QDoubleSpinBox::down-button:hover { background: #F3F4F6; }
            QSpinBox::up-button:pressed, QDoubleSpinBox::up-button:pressed, QSpinBox::down-button:pressed, QDoubleSpinBox::down-button:pressed { background: #EDE9FE; }
            QTableWidget { background: #FFFFFF; border: 1px solid #E5E7EB; border-radius: 10px; gridline-color: #F1F5F9; }
            QHeaderView::section { background: #F3F4F6; color: #1F2937; border: none; padding: 10px 46px 10px 10px; font-weight: 700; }
            QHeaderView::section:hover { background: #EDE9FE; color: #6D28D9; }
            QTableWidget::item { padding: 10px 8px; color: #111827; }
            QScrollBar:vertical { background: #F8FAFC; width: 10px; margin: 2px; border-radius: 5px; }
            QScrollBar::handle:vertical { background: #CBD5E1; min-height: 28px; border-radius: 5px; }
            QScrollBar::handle:vertical:hover { background: #94A3B8; }
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical { height: 0; border: none; background: transparent; }
            QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical { background: transparent; }
            QScrollBar:horizontal { background: #F8FAFC; height: 10px; margin: 2px; border-radius: 5px; }
            QScrollBar::handle:horizontal { background: #CBD5E1; min-width: 28px; border-radius: 5px; }
            QScrollBar::handle:horizontal:hover { background: #94A3B8; }
            QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal { width: 0; border: none; background: transparent; }
            QScrollBar::add-page:horizontal, QScrollBar::sub-page:horizontal { background: transparent; }
            QTableWidget::item:selected { background: #F3E8FF; color: #111827; }
            #modalTitleBar { background: transparent; border: none; max-height: 32px; }
            #modalCaption { color: #111827; font-size: 11pt; font-weight: 800; }
            #infoDialogIcon { background: #DBEAFE; color: #2563EB; border-radius: 17px; font-size: 13pt; font-weight: 800; }
            #errorDialogIcon { background: #FEE2E2; color: #EF4444; border-radius: 17px; font-size: 13pt; font-weight: 800; }
            #modalTitleBar { background: transparent; border: none; }
            #modalCloseButton { background: #F8FAFC; color: #374151; border: 1px solid #E5E7EB; border-radius: 14px; font-size: 10pt; font-weight: 800; min-width: 28px; max-width: 28px; min-height: 28px; max-height: 28px; padding: 0; }
            #modalCloseButton:hover { background: #FEE2E2; color: #EF4444; }
            #modalCloseButton:pressed { background: #FCA5A5; color: white; }
            #toastSuccess { background: #16A34A; color: white; border-radius: 14px; padding: 12px 18px; font-weight: 800; border: 1px solid #15803D; }
            #windowTitleBar { background: #FFFFFF; border-bottom: 1px solid #E5E7EB; }
            #windowCaption { color: #374151; font-weight: 800; }
            #windowControlButton { background: transparent; border: none; border-radius: 10px; font-size: 11pt; font-weight: 800; min-width: 34px; max-width: 34px; min-height: 30px; max-height: 30px; padding: 0; }
            #windowControlButton:hover { background: #F3F4F6; }
            #windowCloseButton { background: transparent; color: #EF4444; border: none; border-radius: 10px; font-size: 11pt; font-weight: 800; min-width: 34px; max-width: 34px; min-height: 30px; max-height: 30px; padding: 0; }
            #windowCloseButton:hover { background: #FEE2E2; color: #DC2626; }
            #refreshButton { background: #FFFFFF; color: #7E22CE; border: 1px solid #E9D5FF; border-radius: 12px; font-size: 13pt; font-weight: 900; min-width: 38px; max-width: 38px; min-height: 34px; max-height: 34px; padding: 0; }
            #refreshButton:hover { background: #F3E8FF; border-color: #8A1CF6; }
            #refreshButton:pressed { background: #E9D5FF; }
            QTabWidget::pane { border: 1px solid #E5E7EB; border-radius: 12px; background: #FFFFFF; top: -1px; }
            QTabBar::tab { background: #F8FAFC; color: #374151; border: 1px solid #E5E7EB; border-bottom: none; border-top-left-radius: 10px; border-top-right-radius: 10px; padding: 8px 14px; margin-right: 4px; min-height: 24px; }
            QTabBar::tab:selected { background: #FFFFFF; color: #8A1CF6; font-weight: 700; }
            QTabBar::tab:hover { background: #F3E8FF; }
            QTabBar::close-button { width: 13px; height: 13px; margin-left: 5px; subcontrol-position: right; }
            QTabBar::close-button:hover { background: #FEE2E2; border-radius: 6px; }
        """.replace("__COMBO_ARROW_ICON__", combo_arrow_icon).replace("__PLUS_ICON__", plus_icon).replace("__MINUS_ICON__", minus_icon))
        root=QWidget(); outer_layout=QVBoxLayout(root); outer_layout.setContentsMargins(0,0,0,0); outer_layout.setSpacing(0)
        self._main_titlebar=QFrame(); self._main_titlebar.setObjectName("windowTitleBar"); self._main_titlebar.setFixedHeight(38); self._main_titlebar.installEventFilter(self)
        tb=QHBoxLayout(self._main_titlebar); tb.setContentsMargins(12,4,8,4); tb.setSpacing(6)
        tb.addWidget(QLabel("Boxio"),1)
        min_btn=QPushButton("–"); min_btn.setObjectName("windowControlButton"); min_btn.setToolTip("Minimizar janela"); min_btn.clicked.connect(self.showMinimized)
        self._maximize_button=QPushButton("□"); self._maximize_button.setObjectName("windowControlButton"); self._maximize_button.setToolTip("Maximizar janela"); self._maximize_button.clicked.connect(self.toggle_maximize_restore)
        close_btn=QPushButton("✕"); close_btn.setObjectName("windowCloseButton"); close_btn.setToolTip("Fechar sistema"); close_btn.clicked.connect(self.close)
        tb.addWidget(min_btn); tb.addWidget(self._maximize_button); tb.addWidget(close_btn)
        self.update_maximize_button_icon()
        body=QWidget(); root_layout=QHBoxLayout(body); root_layout.setContentsMargins(0,0,0,0); root_layout.setSpacing(0)
        outer_layout.addWidget(self._main_titlebar); outer_layout.addWidget(body,1)
        sidebar=QFrame(); sidebar.setObjectName("sidebar"); sidebar.setFixedWidth(235); side=QVBoxLayout(sidebar); side.setContentsMargins(18,22,18,22)
        brand_container = QFrame(); brand_container.setObjectName("brandContainer")
        brand_layout = QHBoxLayout(brand_container); brand_layout.setContentsMargins(0,0,0,0); brand_layout.setSpacing(10)
        logo = QLabel("CE"); logo.setObjectName("logoBox"); logo.setAlignment(Qt.AlignCenter); logo.setToolTip("Espaço reservado para o logo/ícone da empresa. Substitua assets/logo.png para usar o logo real.")
        logo_path = ROOT_DIR / "assets" / "logo.png"
        if logo_path.exists():
            pix = QPixmap(str(logo_path))
            if not pix.isNull():
                logo.setText("")
                logo.setPixmap(pix.scaled(28, 28, Qt.KeepAspectRatio, Qt.SmoothTransformation))
        brand=QLabel("Boxio"); brand.setObjectName("brand"); brand.setWordWrap(True)
        brand_layout.addWidget(logo); brand_layout.addWidget(brand,1)
        side.addWidget(brand_container); side.addSpacing(12)
        self.pages=QStackedWidget()
        navs=[("📊 Dashboard",self.show_dashboard),("📦 Inventário",self.show_inventory),("   ↳ ➕ Novo Produto",self.add_product),("🔁 Movimentações",self.show_movements),("🛒 Compras",self.show_purchases),("⚙️ Cadastros",self.show_registry),("⬆️ Atualizações",self.check_updates_manual)]
        for text,cb in navs:
            b=QPushButton(text); b.setObjectName("navButton"); b.clicked.connect(cb); side.addWidget(b)
        side.addStretch(); root_layout.addWidget(sidebar); root_layout.addWidget(self.pages,1); self.setCentralWidget(root)
        self.toast = ToastNotification(self)
        self.dashboard_page=self.build_dashboard_page(); self.inventory_page=self.build_inventory_page(); self.movements_page=self.build_movements_page(); self.purchases_page=self.build_purchases_page(); self.registry_page=self.build_registry_page()
        for p in [self.dashboard_page,self.inventory_page,self.movements_page,self.purchases_page,self.registry_page]: self.pages.addWidget(p)
        apply_boxio_control_symbols(self)

    def page_shell(self,title,subtitle):
        page=QWidget(); layout=QVBoxLayout(page); layout.setContentsMargins(26,22,26,22); layout.setSpacing(16)
        title_row=QHBoxLayout(); t=QLabel(title); t.setObjectName("pageTitle"); title_row.addWidget(t); title_row.addWidget(HelpIcon("registry_page" if title=="Cadastros e Estrutura" else subtitle)); title_row.addStretch(); refresh_btn=QPushButton("⟳"); refresh_btn.setObjectName("refreshButton"); refresh_btn.setToolTip("Atualizar dados da tela atual"); refresh_btn.clicked.connect(self.force_refresh_current_page); title_row.addWidget(refresh_btn); layout.addLayout(title_row)
        s=QLabel(subtitle); s.setObjectName("pageSubtitle"); layout.addWidget(s); return page,layout

    def build_dashboard_page(self):
        page,layout=self.page_shell("Dashboard","Visão geral do estoque, alertas, compras e atividades recentes."); self.kpi_grid=QGridLayout(); layout.addLayout(self.kpi_grid)
        mid=QHBoxLayout(); recent_card=QFrame(); recent_card.setObjectName("card"); rc=QVBoxLayout(recent_card); rh=QHBoxLayout(); rh.addWidget(QLabel("Atividade recente")); rh.addWidget(HelpIcon("recent_activity")); rh.addStretch(); recent_export=QPushButton("⬇ Excel"); recent_export.setToolTip("Exportar atividades recentes visíveis para Excel"); rh.addWidget(recent_export); rc.addLayout(rh); self.recent_table=QTableWidget(0,6); recent_export.clicked.connect(lambda: export_table_to_excel(self,self.recent_table,"atividades_recentes.xlsx")); self.recent_table.setHorizontalHeaderLabels(["Atividade","Produto","Qtd.","Status","Responsável","Data"]); configure_table(self.recent_table,self.service,"recent_activities"); self.recent_table.itemDoubleClicked.connect(lambda *_: self.open_selected_recent()); rc.addWidget(self.recent_table)
        low_card=QFrame(); low_card.setObjectName("card"); lc=QVBoxLayout(low_card); head=QHBoxLayout(); head.addWidget(QLabel("Alerta de estoque baixo")); head.addWidget(HelpIcon("low_stock")); view=QPushButton("Ver todos"); view.clicked.connect(lambda: self.open_kpi_detail("low")); export_low=QPushButton("⬇ Excel"); export_low.setToolTip("Exportar alerta de estoque baixo para Excel"); head.addStretch(); head.addWidget(export_low); head.addWidget(view); lc.addLayout(head); self.low_table=QTableWidget(0,6); export_low.clicked.connect(lambda: export_table_to_excel(self,self.low_table,"estoque_baixo.xlsx")); self.low_table.setHorizontalHeaderLabels(["Produto","SKU","Estoque","Mínimo","Status","Compra"]); configure_table(self.low_table,self.service,"low_stock"); self.low_table.itemDoubleClicked.connect(lambda *_: self.open_low_table_item()); lc.addWidget(self.low_table)
        mid.addWidget(recent_card,2); mid.addWidget(low_card,1); layout.addLayout(mid); return page

    def build_inventory_page(self):
        page,layout=self.page_shell("Inventário","Gerencie produtos, unidades, estoque, compras em andamento e ações operacionais."); top=QHBoxLayout(); self.search=QLineEdit(); self.search.setPlaceholderText("Buscar por produto, SKU, categoria, marca, status..."); self._inventory_search_timer = QTimer(self); self._inventory_search_timer.setSingleShot(True); self._inventory_search_timer.timeout.connect(self.refresh_inventory); self.search.textChanged.connect(lambda: self._inventory_search_timer.start(180)); self.filter_column=QComboBox(); self.filter_column.addItems(["Todas as colunas","Produto","SKU","Categoria","Marca","Status","Compra"]); self.filter_column.currentIndexChanged.connect(self.refresh_inventory); add=QPushButton("+ Adicionar Produto"); add.setObjectName("primaryButton"); add.clicked.connect(self.add_product); export=QPushButton("⬇ Exportar Excel"); export.setToolTip("Exporta linhas visíveis ou selecionadas para .xlsx"); export.clicked.connect(lambda: export_table_to_excel(self,self.inventory_table,"inventario.xlsx")); top.addWidget(self.search,1); top.addWidget(self.filter_column); top.addWidget(export); top.addWidget(add); layout.addLayout(top)
        self.inventory_page_info=QLabel(""); self.inventory_page_info.setObjectName("hint"); layout.addWidget(self.inventory_page_info)
        self.inventory_table=QTableWidget(0,12); self.inventory_table.setHorizontalHeaderLabels(["Produto","SKU","Status","Categoria","Tipo","Unid.","Qtd. Base","Estoque","Físico","Preço","Compra","ID"]); self.inventory_table.setColumnHidden(11,True); configure_table(self.inventory_table,self.service,"inventory"); self.inventory_table.itemDoubleClicked.connect(lambda *_: self.edit_selected()); layout.addWidget(self.inventory_table,1)
        actions=QHBoxLayout(); edit=QPushButton("Editar"); move=QPushButton("Movimentar"); buy=QPushButton("Solicitar Compra"); delete=QPushButton("Excluir"); delete.setObjectName("dangerButton"); edit.clicked.connect(self.edit_selected); move.clicked.connect(self.move_selected); buy.clicked.connect(self.purchase_selected); delete.clicked.connect(self.delete_selected); actions.addWidget(edit); actions.addWidget(move); actions.addWidget(buy); actions.addWidget(delete); actions.addStretch(); layout.addLayout(actions); return page

    def build_movements_page(self):
        page,layout=self.page_shell("Movimentações de Estoque","Histórico rastreável com unidade, conversão, saldo e responsável."); top=QHBoxLayout(); export=QPushButton("⬇ Exportar Excel"); export.setToolTip("Exporta movimentações visíveis ou selecionadas para .xlsx"); export.clicked.connect(lambda: export_table_to_excel(self,self.mov_table,"movimentacoes.xlsx")); top.addStretch(); top.addWidget(export); layout.addLayout(top); self.mov_table=QTableWidget(0,10); self.mov_table.setHorizontalHeaderLabels(["Produto","SKU","Tipo","Qtd. usada","Unid.","Qtd. convertida","Saldo antes","Saldo restante","Responsável","Data"]); configure_table(self.mov_table,self.service,"movements"); layout.addWidget(self.mov_table); return page

    def build_purchases_page(self):
        page,layout=self.page_shell("Compras e Reposição","Solicitações, aprovação, pedido realizado, recebimento e atualização automática do estoque."); top=QHBoxLayout(); self.purchase_filter=QComboBox(); self.purchase_filter.addItems(["Ativas","Aguardando recebimento","Finalizadas/Canceladas","Todas"]); self.purchase_filter.currentIndexChanged.connect(self.refresh_purchases); new=QPushButton("+ Solicitar Compra"); new.setObjectName("primaryButton"); new.clicked.connect(self.new_purchase); export=QPushButton("⬇ Exportar Excel"); export.setToolTip("Exporta solicitações de compra visíveis ou selecionadas para .xlsx"); export.clicked.connect(lambda: export_table_to_excel(self,self.purchase_table,"compras_reposicao.xlsx")); top.addWidget(self.purchase_filter); top.addStretch(); top.addWidget(export); top.addWidget(new); layout.addLayout(top)
        self.purchase_table=QTableWidget(0,10); self.purchase_table.setHorizontalHeaderLabels(["Item","SKU","Status","Qtd. Solicitada","Recebido","Unid.","Prioridade","Fornecedor","Atualizado em","ID"]); self.purchase_table.setColumnHidden(9,True); configure_table(self.purchase_table,self.service,"purchases"); self.purchase_table.itemDoubleClicked.connect(lambda *_: self.manage_selected_purchase()); layout.addWidget(self.purchase_table,1)
        actions=QHBoxLayout(); manage=QPushButton("Abrir Fluxo"); receive=QPushButton("Receber/Atualizar"); manage.setObjectName("primaryButton"); manage.clicked.connect(self.manage_selected_purchase); receive.clicked.connect(self.manage_selected_purchase); actions.addWidget(manage); actions.addWidget(receive); actions.addStretch(); layout.addLayout(actions); return page

    def build_registry_page(self):
        page,layout=self.page_shell("Cadastros e Estrutura","Gerencie categorias, marcas, fornecedores, locais de estoque e responsáveis."); self.reference_tables={"categories":{"titulo":"Categorias","finalidade":"Classificação dos produtos. Duplo clique mostra itens vinculados."},"brands":{"titulo":"Marcas","finalidade":"Marca controlada sem digitação livre."},"suppliers":{"titulo":"Fornecedores","finalidade":"Base para compras e histórico."},"warehouses":{"titulo":"Estoques/Locais","finalidade":"Locais de armazenamento e estrutura multiestoque."},"responsibles":{"titulo":"Responsáveis","finalidade":"Pessoas responsáveis por movimentações e compras."}}
        self.registry_tabs=QTabWidget(); self.registry_widgets={}
        for table,meta in self.reference_tables.items():
            tab=QWidget(); v=QVBoxLayout(tab); hint_row=QHBoxLayout(); hint=QLabel(meta["finalidade"]); hint.setObjectName("hint"); hint_row.addWidget(hint); hint_row.addWidget(HelpIcon("registry_"+table if table in {"categories","brands","suppliers","warehouses","responsibles"} else meta["finalidade"])); hint_row.addStretch(); v.addLayout(hint_row); tbl=QTableWidget(0,4); tbl.setHorizontalHeaderLabels(["Nome","ID","Origem","Atualizado em"]); tbl.setColumnHidden(1,True); configure_table(tbl,self.service,f"registry_{table}"); tbl.itemDoubleClicked.connect(lambda *_ , t=table: self.reference_double_click(t)); v.addWidget(tbl,1)
            buttons=QHBoxLayout(); add=QPushButton("Adicionar"); edit=QPushButton("Editar"); export=QPushButton("⬇ Exportar Excel"); delete=QPushButton("Excluir"); delete.setObjectName("dangerButton"); add.clicked.connect(lambda _,t=table:self.add_reference(t)); edit.clicked.connect(lambda _,t=table:self.edit_reference(t)); export.clicked.connect(lambda _,tb=tbl,n=table: export_table_to_excel(self,tb,f"cadastro_{n}.xlsx")); delete.clicked.connect(lambda _,t=table:self.delete_reference(t)); buttons.addWidget(add); buttons.addWidget(edit); buttons.addWidget(export); buttons.addWidget(delete); buttons.addStretch(); v.addLayout(buttons); self.registry_tabs.addTab(tab,meta["titulo"]); self.registry_widgets[table]=tbl
        layout.addWidget(self.registry_tabs,1); return page

    def selected_product(self):
        rows=self.inventory_table.selectionModel().selectedRows();
        if not rows: return None
        pid=self.inventory_table.item(rows[0].row(),11).text(); return self.service.get_product(pid)
    def selected_purchase(self):
        rows=self.purchase_table.selectionModel().selectedRows();
        if not rows: return None
        rid=self.purchase_table.item(rows[0].row(),9).text(); return next((r for r in self.service.purchase_requests() if r["id"]==rid),None)
    def current_reference_record(self,table):
        tbl=self.registry_widgets[table]; rows=tbl.selectionModel().selectedRows();
        if not rows: return None
        rid=tbl.item(rows[0].row(),1).text(); return next((r for r in getattr(self.service,table)() if r["id"]==rid),None)

    def show_success(self, message: str):
        if hasattr(self, "toast"):
            self.toast.show_message(message)
        else:
            show_info(self, "Sucesso", message)

    def add_product(self):
        if ProductDialog(self.service,parent=self).exec()==QDialog.Accepted:
            self.refresh_all(); self.pages.setCurrentWidget(self.inventory_page); self.show_success("Produto cadastrado com sucesso.")
    def edit_selected(self):
        p=self.selected_product();
        if not p: show_info(self,"Seleção","Selecione um produto."); return
        if ProductDialog(self.service,p,self).exec()==QDialog.Accepted:
            self.refresh_all(); self.show_success("Produto atualizado com sucesso.")
    def move_selected(self):
        p=self.selected_product();
        if not p: show_info(self,"Seleção","Selecione um produto."); return
        if MovementDialog(self.service,p,self).exec()==QDialog.Accepted:
            self.refresh_all(); self.show_success("Movimentação registrada com sucesso.")
    def purchase_selected(self):
        p=self.selected_product();
        if not p: show_info(self,"Seleção","Selecione um produto."); return
        if PurchaseRequestDialog(self.service,p,self).exec()==QDialog.Accepted:
            self.refresh_all(); self.pages.setCurrentWidget(self.purchases_page); self.show_success("Solicitação de compra criada com sucesso.")
    def delete_selected(self):
        p=self.selected_product();
        if not p: show_info(self,"Seleção","Selecione um produto."); return
        if confirm_action(self,"Excluir produto",f"Excluir/desativar {p['nome']}?", "Esta ação remove o produto da operação ativa, preservando o histórico quando o backend estiver configurado para desativação lógica.", "Excluir/desativar", "Cancelar"):
            try:
                self.service.delete_product(p["id"]); self.refresh_all(); self.show_success("Produto excluído/desativado com sucesso.")
            except Exception as e:
                show_error(self,"Erro",str(e))
    def new_purchase(self):
        if PurchaseRequestDialog(self.service,parent=self).exec()==QDialog.Accepted:
            self.refresh_all(); self.pages.setCurrentWidget(self.purchases_page); self.show_success("Solicitação de compra criada com sucesso.")
    def manage_selected_purchase(self):
        r=self.selected_purchase();
        if not r: show_info(self,"Seleção","Selecione uma solicitação."); return
        d = PurchaseManageDialog(self.service,r,self)
        d.exec(); self.refresh_all()
        if getattr(d, "operation_success", ""):
            self.show_success(d.operation_success)

    def add_reference(self,table):
        d=TextInputDialog("Adicionar cadastro","Nome:",parent=self)
        if d.exec()==QDialog.Accepted and d.input.text().strip():
            try:
                self.service.create_reference(table,d.input.text().strip(),d.description.toPlainText())
                self.refresh_all()
                self.show_success("Cadastro adicionado com sucesso.")
            except Exception as e:
                show_error(self,"Validação",friendly_exception_message(e))
    def edit_reference(self,table):
        rec=self.current_reference_record(table)
        if not rec: show_info(self,"Seleção","Selecione um registro."); return
        d=TextInputDialog("Editar cadastro","Nome:",rec.get("nome",""),rec.get("descricao",""),self)
        if d.exec()==QDialog.Accepted:
            try: self.service.update_reference(table,rec["id"],d.input.text(),d.description.toPlainText()); self.refresh_all()
            except Exception as e: show_error(self,"Validação",friendly_exception_message(e))
    def delete_reference(self,table):
        rec=self.current_reference_record(table)
        if not rec: show_info(self,"Seleção","Selecione um registro."); return
        if confirm_action(self,"Excluir cadastro",f"Excluir/desativar '{rec['nome']}'?", "O registro deixará de aparecer nas listas ativas, mantendo vínculos históricos quando aplicável.", "Excluir/desativar", "Cancelar"):
            try: self.service.delete_reference(table,rec["id"]); self.refresh_all()
            except Exception as e: show_error(self,"Não foi possível excluir",friendly_exception_message(e))
    def reference_double_click(self,table):
        rec=self.current_reference_record(table)
        if not rec: return
        if table=="categories": CategoryDetailDialog(self.service,rec,self).exec(); self.refresh_all()
        else: self.edit_reference(table)

    def force_refresh_current_page(self):
        """Força sincronização remota da tela atual sem travar a UI."""
        if getattr(self, "_refresh_thread", None) and self._refresh_thread.isRunning():
            self.show_success("Atualização já está em andamento.")
            return
        self.show_success("Atualizando dados da tela...")
        self._refresh_thread = DataRefreshThread(self.service, self)
        self._refresh_thread.finished_ok.connect(lambda: (self.refresh_current_page(), self.show_success("Dados atualizados com sucesso.")))
        self._refresh_thread.refresh_error.connect(lambda msg: show_error(self, "Atualização", friendly_exception_message(msg)))
        self._refresh_thread.start()

    def show_dashboard(self): self.pages.setCurrentWidget(self.dashboard_page); self.refresh_dashboard()
    def show_inventory(self): self.pages.setCurrentWidget(self.inventory_page); self.refresh_inventory()
    def show_movements(self): self.pages.setCurrentWidget(self.movements_page); self.refresh_movements()
    def show_purchases(self): self.pages.setCurrentWidget(self.purchases_page); self.refresh_purchases()
    def show_registry(self): self.pages.setCurrentWidget(self.registry_page); self.refresh_registry()

    def refresh_current_page(self):
        current = self.pages.currentWidget() if hasattr(self, "pages") else None
        if current is getattr(self, "inventory_page", None): self.refresh_inventory()
        elif current is getattr(self, "movements_page", None): self.refresh_movements()
        elif current is getattr(self, "purchases_page", None): self.refresh_purchases()
        elif current is getattr(self, "registry_page", None): self.refresh_registry()
        else: self.refresh_dashboard()

    def start_background_sync(self):
        # Sincroniza Neon/PostgreSQL em segundo plano. A UI permanece usando o
        # cache local instantâneo e não bloqueia cliques, digitação ou troca de páginas.
        if hasattr(self.service, "sync_remote_cache_async"):
            try: self.service.sync_remote_cache_async()
            except Exception: pass

    def refresh_all(self): self.refresh_current_page()

    # Atualiza todos os KPIs e tabelas auxiliares do dashboard.
    # A lista de estoque baixo agora mostra também o status operacional do item.
    def refresh_dashboard(self):
        while self.kpi_grid.count():
            item=self.kpi_grid.takeAt(0)
            if item.widget(): item.widget().deleteLater()
        d=self.service.dashboard(); cards=[("Itens em análise",d["itens_analise"],"Pendentes de decisão","Detalhar",lambda:self.open_kpi_detail("analysis"),BLUE),("Estoque baixo",d["baixo_estoque"],"Atenção necessária","Detalhar",lambda:self.open_kpi_detail("low"),RED),("Sem estoque",d["sem_estoque"],"Reposição urgente","Detalhar",lambda:self.open_kpi_detail("out"),RED),("Movimentações",d["movimentacoes"],"Histórico rastreável","Detalhar",lambda:self.open_kpi_detail("movements"),GREEN),("Compras ativas",d["compras_pendentes"],"Solicitações em andamento","Detalhar",lambda:self.open_kpi_detail("purchases"),YELLOW)]
        for i,args in enumerate(cards): self.kpi_grid.addWidget(ActionCard(*args),0,i)
        self.low_table.setSortingEnabled(False); self.low_table.setRowCount(0)
        for p in d.get("itens_baixo", []):
            r=self.low_table.rowCount(); self.low_table.insertRow(r); status="Sem estoque" if float(p.get("estoque_atual",0))<=0 else ("Estoque baixo" if float(p.get("estoque_atual",0))<=float(p.get("estoque_minimo",0)) else "Em estoque"); vals=[p.get("nome",""),p.get("sku",""),fmt_num(p.get("estoque_atual",0)),fmt_num(p.get("estoque_minimo",0)),status,p.get("compra_status","")]
            for c,v in enumerate(vals):
                item=QTableWidgetItem(str(v)); item.setForeground(QColor(TEXT)); item.setToolTip(str(v))
                if c in {4,5} and str(v):
                    set_status_cell(self.low_table,r,c,str(v)); continue
                self.low_table.setItem(r,c,item)
        apply_table_filters(self.low_table)
        self.low_table.setSortingEnabled(True); self.refresh_activities(self.recent_table,8)

    def open_kpi_detail(self,kind):
        if kind=="low": rows=[p for p in self.service.products() if float(p.get("estoque_atual",0))<=float(p.get("estoque_minimo",0))]; DetailDialog("Itens com estoque baixo",self.service,rows,"products",self).exec()
        elif kind=="out": DetailDialog("Itens sem estoque",self.service,self.service.out_stock_products(),"products",self).exec()
        elif kind=="analysis": DetailDialog("Itens em análise",self.service,self.service.analysis_products(),"products",self).exec()
        elif kind=="movements": DetailDialog("Movimentações recentes",self.service,self.service.recent_movements(80),"movements",self).exec()
        elif kind=="purchases": DetailDialog("Compras em andamento",self.service,self.service.purchase_requests(statuses=ACTIVE_PURCHASE_STATUSES),"purchases",self).exec()
        self.refresh_all()
    def open_low_table_item(self): self.open_kpi_detail("low")
    def open_selected_recent(self): self.open_kpi_detail("movements")

    # Recarrega a tabela principal sempre a partir do serviço, garantindo que filtros,
    # status de compra e saldos calculados permaneçam sincronizados com o JSON.
    def refresh_inventory(self):
        query=self.search.text().strip().lower() if hasattr(self,"search") else ""; col_filter=self.filter_column.currentText() if hasattr(self,"filter_column") else "Todas as colunas"
        self.inventory_table.setUpdatesEnabled(False); self.inventory_table.setSortingEnabled(False); self.inventory_table.setRowCount(0)
        products = self.service.products()
        render_limit = 500
        total_matches = 0
        units_cache = self.service.units()
        for p in products:
            cat=self.service.category_name(p.get("categoria_id","")); marca=self.service.brand_name(p.get("marca_id","")) or p.get("marca",""); status="Sem estoque" if float(p.get("estoque_atual",0))<=0 else ("Estoque baixo" if float(p.get("estoque_atual",0))<=float(p.get("estoque_minimo",0)) else "Em estoque"); compra=p.get("compra_status", "")
            fmap={"Produto":p.get("nome",""),"SKU":p.get("sku",""),"Categoria":cat,"Marca":marca,"Status":status,"Compra":compra}; hay=(fmap.get(col_filter) if col_filter!="Todas as colunas" else " ".join([p.get("nome",""),p.get("sku",""),cat,marca,status,compra])).lower()
            if query and query not in hay: continue
            total_matches += 1
            if self.inventory_table.rowCount() >= render_limit:
                continue
            unit=p.get("unidade_medida",""); fisico=physical_total(p.get("estoque_atual",0),p.get("quantidade_base",1),unit,units_cache); vals=[p.get("nome",""),p.get("sku",""),status,cat,p.get("tipo_material",""),unit,fmt_num(p.get("quantidade_base",1)),fmt_num(p.get("estoque_atual",0)),f"{fmt_num(fisico)} {unit}",money(p.get("preco_custo",0)),compra,p["id"]]
            r=self.inventory_table.rowCount(); self.inventory_table.insertRow(r)
            for c,v in enumerate(vals):
                item=QTableWidgetItem(str(v)); item.setForeground(QColor(TEXT)); item.setToolTip(str(v))
                if c in {2,10} and str(v):
                    set_status_cell(self.inventory_table,r,c,str(v)); continue
                self.inventory_table.setItem(r,c,item)
        if hasattr(self, "inventory_page_info"):
            extra = "" if total_matches <= render_limit else f" Mostrando os primeiros {render_limit}; use a busca/filtro para refinar."
            self.inventory_page_info.setText(f"{total_matches} item(ns) encontrado(s)." + extra)
        apply_table_filters(self.inventory_table)
        self.inventory_table.setSortingEnabled(True); self.inventory_table.setUpdatesEnabled(True)


    def refresh_activities(self, table=None, limit=50):
        target=table or self.recent_table
        target.setSortingEnabled(False); target.setRowCount(0)
        for a in self.service.recent_activities(limit):
            vals=[a.get("tipo",""), a.get("produto_nome",""), fmt_num(a.get("quantidade",0)) if a.get("quantidade","")!="" else "", a.get("status",""), a.get("responsavel",""), a.get("criado_em","")]
            r=target.rowCount(); target.insertRow(r)
            for c,v in enumerate(vals):
                if c==3: set_status_cell(target,r,c,str(v))
                else:
                    item=QTableWidgetItem(str(v)); item.setForeground(QColor(TEXT)); item.setToolTip(str(v)); target.setItem(r,c,item)
        apply_table_filters(target)
        target.setSortingEnabled(True)

    def refresh_movements(self,table=None,limit=500):
        target=table or self.mov_table; target.setUpdatesEnabled(False); target.setSortingEnabled(False); target.setRowCount(0)
        for m in self.service.recent_movements(limit):
            vals=[m.get("produto_nome",""),m.get("sku",""),m.get("tipo",""),fmt_num(m.get("quantidade",0)),m.get("unidade_utilizada",""),fmt_num(m.get("quantidade_convertida",0)),fmt_num(m.get("saldo_anterior",0)),fmt_num(m.get("saldo_restante",0)),m.get("responsavel",""),m.get("criado_em","")]
            if target.columnCount()==6: vals=[vals[0],vals[1],vals[2],vals[3],vals[7],vals[9]]
            r=target.rowCount(); target.insertRow(r)
            for c,v in enumerate(vals): item=QTableWidgetItem(str(v)); item.setForeground(QColor(TEXT)); item.setToolTip(str(v)); target.setItem(r,c,item)
        target.setSortingEnabled(True); target.setUpdatesEnabled(True)
        apply_table_filters(target)

    def refresh_purchases(self):
        if not hasattr(self,"purchase_table"): return
        mode=self.purchase_filter.currentText(); statuses=None
        if mode=="Ativas": statuses=ACTIVE_PURCHASE_STATUSES
        elif mode=="Aguardando recebimento": statuses={"Pedido Realizado","Compra Parcial Recebida"}
        elif mode=="Finalizadas/Canceladas": statuses=FINAL_PURCHASE_STATUSES | {"Rejeitado"}
        self.purchase_table.setUpdatesEnabled(False); self.purchase_table.setSortingEnabled(False); self.purchase_table.setRowCount(0)
        for r in self.service.purchase_requests(statuses=statuses):
            vals=[r.get("produto_nome",""),r.get("sku",""),r.get("status",""),fmt_num(r.get("quantidade_solicitada",0)),fmt_num(r.get("quantidade_recebida",0)),r.get("unidade_medida",""),r.get("prioridade",""),r.get("fornecedor",""),r.get("atualizado_em",r.get("criado_em","")),r["id"]]
            row=self.purchase_table.rowCount(); self.purchase_table.insertRow(row)
            for c,v in enumerate(vals):
                item=QTableWidgetItem(str(v)); item.setForeground(QColor(TEXT)); item.setToolTip(str(v))
                if c==2 and str(v):
                    set_status_cell(self.purchase_table,row,c,str(v)); continue
                self.purchase_table.setItem(row,c,item)
        apply_table_filters(self.purchase_table)
        self.purchase_table.setSortingEnabled(True); self.purchase_table.setUpdatesEnabled(True)

    def refresh_registry(self):
        if not hasattr(self,"registry_widgets"): return
        for table,tbl in self.registry_widgets.items():
            tbl.setSortingEnabled(False); tbl.setRowCount(0)
            for rec in getattr(self.service,table)():
                r=tbl.rowCount(); tbl.insertRow(r); vals=[rec.get("nome",""),rec.get("id",""),rec.get("origem","usuario"),rec.get("atualizado_em",rec.get("criado_em",""))]
                for c,v in enumerate(vals): item=QTableWidgetItem(str(v)); item.setForeground(QColor(TEXT)); item.setToolTip(str(v)); tbl.setItem(r,c,item)
            tbl.setSortingEnabled(True)
            apply_table_filters(tbl)
